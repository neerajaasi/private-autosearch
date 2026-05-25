import os
import time
import random
import re
import requests
import yaml
import pandas as pd
from bs4 import BeautifulSoup
from urllib.parse import quote
from datetime import datetime, timedelta, date
from openpyxl import Workbook, load_workbook

# ------------------------------------------------------
# CONSTANTS
# ------------------------------------------------------
RESULTS_PER_PAGE = 25

# ------------------------------------------------------
# CONFIG PATH
# ------------------------------------------------------
CURRENT_DIR = os.path.dirname(os.path.abspath(__file__))
JOBSEARCH_ROOT = os.path.abspath(os.path.join(CURRENT_DIR, ".."))
CONFIG_PATH = os.path.join(JOBSEARCH_ROOT, "config", "linkedinconfig-cad.yaml")

# ------------------------------------------------------
# LOAD CONFIG
# ------------------------------------------------------
def load_config():
    with open(CONFIG_PATH, "r", encoding="utf-8") as f:
        return yaml.safe_load(f)

# ------------------------------------------------------
# EXCEL FILE CREATION
# ------------------------------------------------------
def generate_output_file(output_root):
    today = datetime.now().strftime("%Y-%m-%d")
    out_dir = os.path.join(JOBSEARCH_ROOT, output_root)
    os.makedirs(out_dir, exist_ok=True)

    base = f"LinkedIn_Jobs_{today}.xlsx"
    file_path = os.path.join(out_dir, base)

    counter = 1
    while os.path.exists(file_path):
        file_path = os.path.join(out_dir, f"LinkedIn_Jobs_{today}_run{counter}.xlsx")
        counter += 1

    return file_path

def initialize_excel_file(filepath):
    if not os.path.exists(filepath):
        wb = Workbook()
        wb.active.title = "Sheet1"
        wb.save(filepath)

# ------------------------------------------------------
# BUILD SEARCH URL
# ------------------------------------------------------
def build_url(job_title, region, posted_hours, job_type, start):
    posted_param = f"r{posted_hours * 3600}"
    is_remote = "remote" in region.lower()

    base_url = "https://ca.linkedin.com/jobs/search/?"

    if is_remote:
        # Canada-focused remote search
        quoted_title = quote(job_title + " remote canada")

        return (
            f"{base_url}"
            f"keywords={quoted_title}&"
            "&f_WT=2"
            f"&f_TPR={posted_param}"
            f"&f_JT={job_type}"
            f"&start={start}"
        )
    else:
        # Canada onsite/hybrid jobs
        quoted_title = quote(job_title)

        return (
            f"{base_url}"
            f"keywords={quoted_title}&"
            "location=Canada&"
            "geoId=101174742&"
            f"&f_TPR={posted_param}"
            f"&f_JT={job_type}"
            f"&start={start}"
        )

# ------------------------------------------------------
# FETCH HTML
# ------------------------------------------------------
def fetch_html(url):
    headers = {
        "User-Agent": (
            "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
            "AppleWebKit/537.36 (KHTML, like Gecko) "
            "Chrome/122.0.0.0 Safari/537.36"
        )
    }
    try:
        r = requests.get(url, headers=headers, timeout=20)
        return r.text if r.status_code == 200 else None
    except:
        return None

# ------------------------------------------------------
# POSTED TIME NORMALIZATION
# ------------------------------------------------------
def normalize_posted(text):
    if not text:
        return ""
    text = text.lower().strip()
    today = date.today()

    try:
        if "minute" in text or "hour" in text:
            return today.isoformat()
        if "day" in text:
            return (today - timedelta(days=int(text.split()[0]))).isoformat()
    except:
        pass

    return text

# ------------------------------------------------------
# PARSE SEARCH RESULTS
# ------------------------------------------------------
def parse_search(html, job_type, remote_mode):
    soup = BeautifulSoup(html, "html.parser")
    cards = soup.select("div.base-card")

    rows = []

    for c in cards:
        title_el = c.select_one("h3")
        company_el = c.select_one("h4")
        location_el = c.select_one("span.job-search-card__location")
        link_el = c.find("a", class_="base-card__full-link")
        time_el = c.select_one("time")

        posted_raw = time_el.text.strip() if time_el else ""
        posted = normalize_posted(posted_raw)

        loc = location_el.text.strip() if location_el else ""
        if remote_mode:
            loc = "Remote"

        job_url = link_el["href"].split("?")[0] if link_el else ""

        rows.append({
            "title": title_el.text.strip() if title_el else "",
            "company": company_el.text.strip() if company_el else "",
            "location": loc,
            "job_type": job_type,
            "posted": posted,
            "posted_raw": posted_raw,
            "url": job_url,
        })

    return rows

# ------------------------------------------------------
# SAFE SHEET APPEND
# ------------------------------------------------------
def append_sheet(filepath, sheet_name, df):
    wb = load_workbook(filepath)

    if "Sheet1" in wb.sheetnames and len(wb.sheetnames) == 1:
        wb["Sheet1"].title = sheet_name
        wb.save(filepath)
        with pd.ExcelWriter(filepath, engine="openpyxl", mode="a", if_sheet_exists="replace") as writer:
            df.to_excel(writer, sheet_name=sheet_name, index=False)
        return

    with pd.ExcelWriter(filepath, engine="openpyxl", mode="a") as writer:
        df.to_excel(writer, sheet_name=sheet_name, index=False)

# ------------------------------------------------------
# SAVE SHEET
# ------------------------------------------------------
def save_sheet(job_title, records, filepath):
    if not records:
        print(f"[SKIP] No results → {job_title}")
        return

    df = pd.DataFrame(records).drop_duplicates(subset=["url"])

    if df.empty:
        print(f"[SKIP] Empty after dedupe → {job_title}")
        return

    append_sheet(filepath, job_title[:31], df)
    print(f"[SAVED] {job_title} ({len(df)} rows)")

# ------------------------------------------------------
# MAIN
# ------------------------------------------------------
def run():
    cfg = load_config()

    job_titles = cfg["job_titles"]
    posted_hours = cfg["filters"]["posted_time_hours"]
    job_types = cfg["filters"]["job_type"]
    output_root = cfg.get("output_root", "results")
    max_pages = cfg.get("max_pages", 10)
    regions = ["Canada", "remote"]

    print(f"[CONFIG] posted_hours={posted_hours}, max_pages={max_pages}")

    excel_path = generate_output_file(output_root)
    initialize_excel_file(excel_path)

    for job_title in job_titles:
        combined = []
        print(f"\n=========== {job_title} ===========")

        for job_type in job_types:
            for region in regions:
                print(f"[SEARCH] {job_type} | {region}")

                start = 0
                page = 1

                while page <= max_pages:
                    print(f"  [PAGE] {page}")

                    url = build_url(
                        job_title,
                        region,
                        posted_hours,
                        job_type,
                        start
                    )

                    html = fetch_html(url)
                    if not html:
                        break

                    results = parse_search(
                        html,
                        job_type,
                        region.lower() == "remote"
                    )

                    print(f"    → jobs found: {len(results)}")

                    if not results:
                        break

                    combined.extend(results)

                    if len(results) < RESULTS_PER_PAGE:
                        break

                    start += RESULTS_PER_PAGE
                    page += 1
                    time.sleep(random.uniform(0.6, 1.1))

        save_sheet(job_title, combined, excel_path)

# ------------------------------------------------------
# ENTRY POINT
# ------------------------------------------------------
if __name__ == "__main__":
    run()
