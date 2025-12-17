import os
import time
import random
import re
import requests
import yaml
import pandas as pd
from bs4 import BeautifulSoup
from urllib.parse import quote
from datetime import datetime, timedelta, date
from openpyxl import Workbook, load_workbook


# ------------------------------------------------------
# CONFIG PATH
# ------------------------------------------------------
CURRENT_DIR = os.path.dirname(os.path.abspath(__file__))
JOBSEARCH_ROOT = os.path.abspath(os.path.join(CURRENT_DIR, ".."))
CONFIG_PATH = os.path.join(JOBSEARCH_ROOT, "config", "linkedinconfig.yaml")


# ------------------------------------------------------
# LOAD CONFIG
# ------------------------------------------------------
def load_config():
    with open(CONFIG_PATH, "r", encoding="utf-8") as f:
        return yaml.safe_load(f)


# ------------------------------------------------------
# EXCEL FILE CREATION
# ------------------------------------------------------
def generate_output_file(output_root):
    today = datetime.now().strftime("%Y-%m-%d")
    out_dir = os.path.join(JOBSEARCH_ROOT, output_root, "linkedin")
    os.makedirs(out_dir, exist_ok=True)

    base = f"LinkedIn_Jobs_{today}.xlsx"
    file_path = os.path.join(out_dir, base)

    counter = 2
    while os.path.exists(file_path):
        file_path = os.path.join(out_dir, f"LinkedIn_Jobs_{today}_run{counter}.xlsx")
        counter += 1

    return file_path


def initialize_excel_file(filepath):
    if not os.path.exists(filepath):
        wb = Workbook()
        wb.active.title = "Sheet1"
        wb.save(filepath)


# ------------------------------------------------------
# BUILD SEARCH URL
# ------------------------------------------------------
def build_url(job_title, region, posted_hours, job_type):
    posted_param = f"r{posted_hours * 3600}"
    remote_filter = "&f_WT=2" if region.lower() == "remote" else ""

    return (
        "https://www.linkedin.com/jobs/search/?"
        f"keywords={quote(job_title)}&"
        "location=United%20States&"
        f"f_JT={job_type}"
        f"{remote_filter}"
        f"&f_TPR={posted_param}"
    )


# ------------------------------------------------------
# FETCH HTML
# ------------------------------------------------------
def fetch_html(url):
    headers = {
        "User-Agent": (
            "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
            "AppleWebKit/537.36 (KHTML, like Gecko) "
            "Chrome/122.0.0.0 Safari/537.36"
        )
    }
    try:
        r = requests.get(url, headers=headers, timeout=15)
        return r.text if r.status_code == 200 else None
    except:
        return None


# ------------------------------------------------------
# POSTED TIME NORMALIZATION
# ------------------------------------------------------
def normalize_posted(text):
    if not text:
        return ""

    text = text.lower().strip()
    today = date.today()

    try:
        if "hour" in text or "minute" in text:
            return today.isoformat()
        if "day" in text:
            return (today - timedelta(days=int(text.split()[0]))).isoformat()
        if "week" in text:
            return (today - timedelta(days=7 * int(text.split()[0]))).isoformat()
    except:
        pass

    return text


# ------------------------------------------------------
# RATE EXTRACTION (OPTIONAL)
# ------------------------------------------------------
RATE_REGEX = re.compile(
    r'(\$[\d,]+(?:\.\d+)?\s*(?:-|to)?\s*\$?[\d,]+(?:\.\d+)?\s*'
    r'(?:/hr|per hour|hourly|/year|per year|annually|k))',
    re.IGNORECASE
)


def extract_rate(text):
    if not text:
        return ""
    m = RATE_REGEX.search(text)
    return m.group(1) if m else ""


def fetch_job_description(url):
    html = fetch_html(url)
    if not html:
        return ""

    soup = BeautifulSoup(html, "html.parser")
    desc = soup.select_one("div.show-more-less-html__markup")
    return desc.get_text(" ", strip=True) if desc else ""


# ------------------------------------------------------
# PARSE SEARCH RESULTS
# ------------------------------------------------------
def parse_search(html, job_type, remote_mode, enable_rate_scrape):
    soup = BeautifulSoup(html, "html.parser")
    cards = soup.select("div.base-card")

    rows = []

    for c in cards:
        title_el = c.select_one("h3")
        company_el = c.select_one("h4")
        location_el = c.select_one("span.job-search-card__location")
        link_el = c.find("a", class_="base-card__full-link")

        time_el = (
                c.select_one("time") or
                c.select_one("span.job-search-card__listdate")
        )

        posted_raw = time_el.text.strip() if time_el else ""
        posted = normalize_posted(posted_raw)

        loc = location_el.text.strip() if location_el else ""
        if remote_mode:
            loc = "Remote"

        job_url = link_el["href"].split("?")[0] if link_el else ""

        rate = ""
        if enable_rate_scrape and job_url:
            desc_text = fetch_job_description(job_url)
            rate = extract_rate(desc_text)
            time.sleep(random.uniform(0.4, 0.7))

        rows.append({
            "title": title_el.text.strip() if title_el else "",
            "company": company_el.text.strip() if company_el else "",
            "location": loc,
            "job_type": job_type,
            "posted": posted,
            "posted_raw": posted_raw,
            "rate": rate,
            "url": job_url,
        })

    return rows


# ------------------------------------------------------
# SAFE SHEET APPEND
# ------------------------------------------------------
def append_sheet(filepath, sheet_name, df):
    wb = load_workbook(filepath)

    if "Sheet1" in wb.sheetnames and len(wb.sheetnames) == 1:
        wb["Sheet1"].title = sheet_name
        wb.save(filepath)
        with pd.ExcelWriter(filepath, engine="openpyxl", mode="a", if_sheet_exists="replace") as writer:
            df.to_excel(writer, sheet_name=sheet_name, index=False)
        return

    with pd.ExcelWriter(filepath, engine="openpyxl", mode="a") as writer:
        df.to_excel(writer, sheet_name=sheet_name, index=False)


# ------------------------------------------------------
# SAVE SHEET
# ------------------------------------------------------
def save_sheet(job_title, records, filepath):
    if not records:
        print(f"[SKIP] No results → {job_title}")
        return

    df = pd.DataFrame(records).drop_duplicates(subset=["url"])

    if df.empty:
        print(f"[SKIP] Empty after dedupe → {job_title}")
        return

    append_sheet(filepath, job_title[:31], df)
    print(f"[SAVED] {job_title} ({len(df)} rows)")


# ------------------------------------------------------
# MAIN
# ------------------------------------------------------
def run():
    cfg = load_config()

    job_titles = cfg["job_titles"]
    posted_hours = cfg["filters"]["posted_time_hours"]
    job_types = cfg["filters"]["job_type"]
    output_root = cfg.get("output_root", "results")
    enable_rate_scrape = cfg.get("features", {}).get("enable_rate_scrape", False)

    regions = ["United States", "Remote"]

    excel_path = generate_output_file(output_root)
    initialize_excel_file(excel_path)

    for job_title in job_titles:
        combined = []
        print(f"\n=========== {job_title} ===========")

        for job_type in job_types:
            for region in regions:
                print(f"[SEARCH] {job_type} | {region}")

                url = build_url(job_title, region, posted_hours, job_type)
                html = fetch_html(url)

                if not html:
                    continue

                results = parse_search(
                    html,
                    job_type,
                    region.lower() == "remote",
                    enable_rate_scrape
                )

                combined.extend(results)
                time.sleep(random.uniform(0.25, 0.5))

        save_sheet(job_title, combined, excel_path)


if __name__ == "__main__":
    run()