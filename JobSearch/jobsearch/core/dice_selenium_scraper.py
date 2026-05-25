import yaml
import re
from pathlib import Path
from time import sleep

from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.chrome.service import Service as ChromeService
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC

from openpyxl import Workbook, load_workbook

from dice_locators import DiceLocators as L

# ------------- CONFIG -------------

BASE_DIR = Path(__file__).resolve().parents[2]
CONFIG_FILE = BASE_DIR / "jobsearch" / "config" / "diceconfig.yaml"
BASE_URL = "https://www.dice.com/jobs"

with open(CONFIG_FILE, "r") as f:
    config = yaml.safe_load(f)

OUTPUT_ROOT = BASE_DIR / config["output_root"]
OUTPUT_ROOT.mkdir(exist_ok=True)
OUTPUT_FILE = OUTPUT_ROOT / "dice_jobs_listitems.xlsx"

COLUMNS = [
    "title",
    "company",
    "location",
    "job_type",
    "posted",
    "posted_raw",
    "rate",
    "url",
]


# ------------- EXCEL HELPERS -------------

def init_workbook(file_path: Path):
    if not file_path.exists():
        wb = Workbook()
        ws = wb.active
        ws.title = "Init"
        wb.save(file_path)


def get_or_create_sheet(wb, sheet_name: str):
    if sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
    else:
        ws = wb.create_sheet(title=sheet_name)
        ws.append(COLUMNS)
    return ws


def append_job_to_excel(job_title: str, job_data: dict):
    init_workbook(OUTPUT_FILE)
    wb = load_workbook(OUTPUT_FILE)

    if "Init" in wb.sheetnames and len(wb.sheetnames) > 1:
        init_ws = wb["Init"]
        if init_ws.max_row == 1 and init_ws.max_column == 1 and init_ws["A1"].value is None:
            wb.remove(init_ws)

    sheet_name = re.sub(r"[:\\\/\?\*\[\]]", "", job_title)[:31]
    ws = get_or_create_sheet(wb, sheet_name)

    row = [job_data.get(col, "") for col in COLUMNS]
    ws.append(row)

    wb.save(OUTPUT_FILE)


# ------------- DICE UTILITIES (SELENIUM) -------------

def accept_cookies_if_present(driver):
    try:
        btn = WebDriverWait(driver, 5).until(
            EC.element_to_be_clickable((By.XPATH, L.COOKIES_ALLOW_ALL_BUTTON))
        )
        btn.click()
    except Exception:
        pass


def ensure_location_united_states(driver):
    loc_input = WebDriverWait(driver, 10).until(
        EC.element_to_be_clickable((By.XPATH, L.LOCATION_INPUT))
    )
    loc_input.click()
    loc_input.send_keys(Keys.CONTROL, "a")
    loc_input.send_keys(Keys.BACKSPACE)
    loc_input.send_keys("United States")
    driver.find_element(By.XPATH, L.SEARCH).click()

    try:
        options = WebDriverWait(driver, 8).until(
            EC.presence_of_all_elements_located(
                (By.XPATH, "//li[@role='option'] | //div[@role='option']")
            )
        )
        for opt in options:
            text = opt.text.strip().lower()
            if opt.text.strip().lower().startswith("united states"):
                opt.click()
                break
    except Exception:
        pass


def open_filters(driver):
    btn = WebDriverWait(driver, 10).until(
        EC.element_to_be_clickable((By.XPATH, L.ALL_FILTERS_BUTTON))
    )
    btn.click()
    WebDriverWait(driver, 15).until(
        EC.visibility_of_element_located((By.XPATH, L.FILTER_DRAWER))
    )


def click_filter_label_if_exists(driver, text):
    try:
        label = WebDriverWait(driver, 3).until(
            EC.element_to_be_clickable(
                (By.XPATH, L.FILTER_LABEL.format(label=text))
            )
        )
        driver.execute_script("arguments[0].scrollIntoView(true);", label)
        label.click()
    except Exception:
        print(f"[WARN] filter option not found: {text}")


def apply_all_filters(driver, filters_cfg):
    WebDriverWait(driver, 15).until(
        EC.presence_of_element_located((By.XPATH, L.FILTER_DRAWER))
    )

    # Expand sections
    for section in ["Work Setting", "Employment Type", "Employer Type", "Posted date"]:
        try:
            btn = driver.find_element(
                By.XPATH, L.FILTER_SECTION_BUTTON.format(section=section)
            )
            driver.execute_script("arguments[0].click();", btn)
        except Exception:
            pass

    # Posted date
    if filters_cfg.get("posted_time_hours") == 24:
        click_filter_label_if_exists(driver, "Today")

    # Work setting
    for ws_value in filters_cfg.get("work_setting", []):
        click_filter_label_if_exists(driver, ws_value)

    # Employment type
    for et_value in filters_cfg.get("employment_type", []):
        click_filter_label_if_exists(driver, et_value)

    # Employer type
    for emp_value in filters_cfg.get("employer_type", []):
        click_filter_label_if_exists(driver, emp_value)

    # Apply
    try:
        apply_btn = driver.find_element(By.XPATH, L.APPLY_FILTERS_BUTTON)
        apply_btn.click()
    except Exception:
        pass

    # Wait for list items
    try:
        WebDriverWait(driver, 20).until(
            EC.presence_of_element_located((By.XPATH, L.LIST_ITEM))
        )
    except Exception:
        print("[WARN] No list items after applying filters")


# ------------- LIST ITEMS & DETAIL PAGES -------------

def parse_title_from_aria(aria_label: str) -> str:
    if not aria_label:
        return ""
    prefix = "View Details for "
    if aria_label.startswith(prefix):
        aria_label = aria_label[len(prefix):]

    m = re.match(r"^(.*)\s*\([^)]*\)\s*$", aria_label)
    if m:
        return m.group(1).strip()
    return aria_label.strip()


def fetch_jobs_from_listitems(driver):
    jobs = []
    list_items = driver.find_elements(By.XPATH, L.LIST_ITEM)
    print(f"[INFO] Found {len(list_items)} list items")

    for li in list_items:
        try:
            link = li.find_element(By.XPATH, L.LIST_ITEM_LINK)
        except Exception:
            continue

        aria_label = link.get_attribute("aria-label")
        href = link.get_attribute("href")
        if not href:
            continue

        if href.startswith("/"):
            href = "https://www.dice.com" + href

        title = parse_title_from_aria(aria_label)

        jobs.append(
            {
                "title": title,
                "url": href,
            }
        )

    return jobs


def extract_job_details(driver, base_fields):
    url = base_fields["url"]
    print(f"[INFO] Extracting details: {url}")

    job_data = {
        "title": base_fields.get("title", ""),
        "company": "",
        "location": "",
        "job_type": "",
        "posted": "",
        "posted_raw": "",
        "rate": "",
        "url": url,
    }

    try:
        driver.get(url)
        WebDriverWait(driver, 15).until(
            EC.presence_of_element_located((By.XPATH, L.BODY_TAG))
        )
        sleep(1)

        # Title override
        try:
            title_elem = driver.find_element(By.XPATH, L.TITLE)
            title_text = title_elem.text.strip()
            if title_text:
                job_data["title"] = title_text
        except Exception:
            pass

        # Company
        try:
            company_elem = driver.find_element(By.XPATH, L.COMPANY)
            job_data["company"] = company_elem.text.strip()
        except Exception:
            pass

        # Location
        try:
            location_elem = driver.find_element(By.XPATH, L.LOCATION)
            job_data["location"] = location_elem.text.strip()
        except Exception:
            pass

        # Job type
        try:
            job_type_elem = driver.find_element(By.XPATH, L.JOB_TYPE)
            job_data["job_type"] = job_type_elem.text.strip()
        except Exception:
            pass

        # Posted
        try:
            posted_elem = driver.find_element(By.XPATH, L.POSTED)
            posted_text = posted_elem.text.strip()
            job_data["posted_raw"] = posted_text

            posted_match = re.search(r"Posted\s+(.+)", posted_text, re.IGNORECASE)
            if posted_match:
                job_data["posted"] = posted_match.group(1).strip()
            else:
                job_data["posted"] = posted_text
        except Exception:
            pass

        # Rate
        try:
            rate_elem = driver.find_element(By.XPATH, L.RATE)
            job_data["rate"] = rate_elem.text.strip()
        except Exception:
            pass

    except Exception as e:
        print(f"[ERROR] Failed to extract {url}: {e}")

    return job_data


# ------------- MAIN -------------

def main():
    service = ChromeService()  # or ChromeService('/path/to/chromedriver')
    options = webdriver.ChromeOptions()
    options.add_argument("--start-maximized")

    driver = webdriver.Chrome(service=service, options=options)
    driver.get(BASE_URL)

    WebDriverWait(driver, 30).until(
        EC.presence_of_element_located((By.XPATH, L.BODY_TAG))
    )

    accept_cookies_if_present(driver)

    for job_title in config["job_titles"]:
        print(f"\n[INFO] ===== Job title: {job_title} =====")

        job_input = WebDriverWait(driver, 15).until(
            EC.element_to_be_clickable((By.XPATH, L.JOB_TITLE_INPUT))
        )
        job_input.clear()
        job_input.send_keys(job_title)
        job_input.send_keys(Keys.ENTER)

        WebDriverWait(driver, 20).until(
            EC.presence_of_element_located((By.XPATH, L.BODY_TAG))
        )
        sleep(1)

        ensure_location_united_states(driver)
        open_filters(driver)
        apply_all_filters(driver, config["filters"])

        base_jobs = fetch_jobs_from_listitems(driver)
        print(f"[INFO] Total jobs (listitems) for '{job_title}': {len(base_jobs)}")

        for base in base_jobs:
            details = extract_job_details(driver, base)
            append_job_to_excel(job_title, details)

        driver.get(BASE_URL)
        WebDriverWait(driver, 20).until(
            EC.presence_of_element_located((By.XPATH, L.BODY_TAG))
        )

    input("Press ENTER to close browser...")
    driver.quit()


if __name__ == "__main__":
    main()
