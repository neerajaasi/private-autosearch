import yaml
from pathlib import Path
from playwright.sync_api import sync_playwright, TimeoutError as PlaywrightTimeoutError
import re
from openpyxl import Workbook, load_workbook
from datetime import datetime, timedelta, date
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC

# ------------- CONFIG -------------

BASE_DIR = Path(__file__).resolve().parents[2]
CONFIG_FILE = BASE_DIR / "jobsearch" / "config" / "diceconfig.yaml"
BASE_URL = "https://www.dice.com/jobs"

with open(CONFIG_FILE, "r") as f:
    config = yaml.safe_load(f)

OUTPUT_ROOT = BASE_DIR / "jobsearch" / config["output_root"]
OUTPUT_ROOT.mkdir(exist_ok=True)
today = datetime.now().strftime("%Y-%m-%d")

BASE_OUTPUT_FILE = OUTPUT_ROOT / f"dice_jobs_listitems_{today}.xlsx"

COLUMNS = [
    "title",
    "company",
    "location",
    "job_type",
    "posted",
    "rate",
    "url",
]

def accept_cookies_if_present(page):
    try:
        # Wait briefly in case popup loads late
        page.wait_for_timeout(1500)

        # Main page
        if page.locator("a.cmpboxbtnyes").count() > 0:
            page.locator("a.cmpboxbtnyes").first.click(force=True)
            print("[INFO] Clicked 'Allow all'")
            return

        # Iframe fallback
        for frame in page.frames:
            if frame.locator("a.cmpboxbtnyes").count() > 0:
                frame.locator("a.cmpboxbtnyes").first.click(force=True)
                print("[INFO] Clicked 'Allow all' (iframe)")
                return

        print("[INFO] Cookie popup not found")

    except Exception as e:
        print(f"[WARN] Cookie handling error: {e}")

# ------------- EXCEL HELPERS -------------

def init_workbook(file_path: Path):
    """
    Ensure OUTPUT_FILE is a valid xlsx.
    If it doesn't exist, create a simple workbook with a visible Init sheet.
    """
    if not file_path.exists():
        wb = Workbook()
        ws = wb.active
        ws.title = "Init"
        wb.save(file_path)

def get_indexed_output_file(base_path: Path) -> Path:
    """
    If file exists, create indexed version:
    file.xlsx
    file_1.xlsx
    file_2.xlsx
    """
    if not base_path.exists():
        return base_path

    stem = base_path.stem
    suffix = base_path.suffix
    parent = base_path.parent

    index = 1
    while True:
        new_file = parent / f"{stem}_{index}{suffix}"
        if not new_file.exists():
            return new_file
        index += 1

OUTPUT_FILE = get_indexed_output_file(BASE_OUTPUT_FILE)

def get_or_create_sheet(wb, sheet_name: str):
    """
    Get sheet by name; if not exists, create and add header row.
    """
    if sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
    else:
        ws = wb.create_sheet(title=sheet_name)
        ws.append(COLUMNS)
    return ws


def append_job_to_excel(job_title: str, job_data: dict):
    """
    Append one job row to the Excel file, in sheet named after job_title.
    Creates workbook/sheet as needed.
    """
    init_workbook(OUTPUT_FILE)

    wb = load_workbook(OUTPUT_FILE)

    # Remove Init sheet if we now have other sheets
    if "Init" in wb.sheetnames and len(wb.sheetnames) > 1:
        init_ws = wb["Init"]
        # If Init is effectively empty, safe to remove
        if init_ws.max_row == 1 and init_ws.max_column == 1 and init_ws["A1"].value is None:
            wb.remove(init_ws)

    sheet_name = re.sub(r"[:\\\/\?\*\[\]]", "", job_title)[:31]
    ws = get_or_create_sheet(wb, sheet_name)

    row = [job_data.get(col, "") for col in COLUMNS]
    ws.append(row)

    wb.save(OUTPUT_FILE)


def ensure_location_united_states(page):
    page.click("input[placeholder*='Location']")
    page.keyboard.press("Control+A")
    page.keyboard.press("Backspace")
    page.fill("input[placeholder*='Location']", "United States")

    page.wait_for_selector("li[role='option'], div[role='option']", timeout=8000)
    options = page.locator("li[role='option'], div[role='option']")
    count = options.count()
    for i in range(count):
        text = options.nth(i).inner_text()
        if text.strip().lower().startswith("united states"):
            options.nth(i).click()
            break


def open_filters(page):
    page.locator("button:has-text('All Filters')").first.click()
    page.locator("text=Filter Results").first.wait_for(state="visible", timeout=15000)

def apply_all_filters(page, filters_cfg):

    # --------------------------------------------------
    # 1️⃣  Wait for Filter Panel to Open
    # --------------------------------------------------
    page.locator("text=Filter Results").first.wait_for(
        state="visible", timeout=15000
    )

    # Small stabilization wait (React render safety)
    page.wait_for_timeout(1000)

    # --------------------------------------------------
    # 2️⃣  Posted Date (Today = 24h)
    # --------------------------------------------------
    try:
        if filters_cfg.get("posted_time_hours") == 24:
            today = page.locator("label:has-text('Today')").first
            today.wait_for(state="visible", timeout=8000)
            today.click()
    except Exception:
        print("[WARN] Posted date filter not applied")

    # --------------------------------------------------
    # 3️⃣  Work Setting (Remote / Hybrid / On-Site)
    # --------------------------------------------------
    for ws_value in filters_cfg.get("work_setting", []):
        try:
            option = page.locator(f"label:has-text('{ws_value}')").first
            option.wait_for(state="visible", timeout=5000)
            option.click()
        except Exception:
            print(f"[WARN] Work setting not found: {ws_value}")

    # --------------------------------------------------
    # 4️⃣  Employment Type (Full time / Contract)
    # --------------------------------------------------
    for et_value in filters_cfg.get("employment_type", []):
        try:
            option = page.locator(f"label:has-text('{et_value}')").first
            option.wait_for(state="visible", timeout=5000)
            option.click()
        except Exception:
            print(f"[WARN] Employment type not found: {et_value}")

    # --------------------------------------------------
    # 5️⃣  Employer Type (Direct hire / Recruiter)
    # --------------------------------------------------
    for emp_value in filters_cfg.get("employer_type", []):
        try:
            option = page.locator(f"label:has-text('{emp_value}')").first
            option.wait_for(state="visible", timeout=5000)
            option.click()
        except Exception:
            print(f"[WARN] Employer type not found: {emp_value}")

    # --------------------------------------------------
    # 6️⃣  Click Apply
    # --------------------------------------------------
    try:
        apply_btn = page.locator("button:has-text('Apply')").first
        apply_btn.wait_for(state="visible", timeout=8000)
        apply_btn.click()
    except Exception:
        print("[ERROR] Apply button not clickable")

    # --------------------------------------------------
    # 7️⃣  Wait for Results to Reload
    # --------------------------------------------------
    try:
        page.wait_for_selector("//div[@role='listitem']", timeout=20000)
    except Exception:
        print("[WARN] No list items found after applying filters")

def parse_title_from_aria(aria_label: str) -> str:
    """
    "View Details for IT Asset Analyst (23b4...)" -> "IT Asset Analyst"
    """
    if not aria_label:
        return ""
    prefix = "View Details for "
    if aria_label.startswith(prefix):
        aria_label = aria_label[len(prefix):]

    m = re.match(r"^(.*)\s*\([^)]*\)\s*$", aria_label)
    if m:
        return m.group(1).strip()
    return aria_label.strip()


def fetch_jobs_from_listitems(page):
    """
    Use //div[@role='listitem'], get inner <a> aria-label and href,
    parse title, return list of {title, url}.
    """
    jobs = []
    list_items = page.locator("//div[@role='listitem']")
    count = list_items.count()
    print(f"[INFO] Found {count} list items")

    for i in range(count):
        li = list_items.nth(i)
        link = li.locator("a").first

        aria_label = link.get_attribute("aria-label")
        title = ""
        if aria_label:
            match = re.search(r"for\s+(.*?)\s*\(", aria_label)
            if match:
                title = match.group(1).strip()
        href = link.get_attribute("href")
        if not href:
            continue
        # Ensure absolute URL
        if href.startswith("/"):
            href = "https://www.dice.com" + href

        # Remove query parameters
        href = href.split("?")[0]

        print(title)
        print(href)

        if not href:
            continue
        jobs.append(
            {
                "title": title,
                "url": href,
            }
        )

    return jobs


def extract_header_details(page, existing_data=None):

    if existing_data is None:
        existing_data = {}

    job_data = existing_data.copy()

    page.wait_for_load_state("domcontentloaded")
    page.wait_for_timeout(1500)

    header = page.locator("[data-testid='job-detail-header-card']")
    if header.count() == 0:
        return job_data

    header = header.first

    # ---------------- Title ----------------
    try:
        title = header.locator("h1").inner_text().strip()
        if title:
            job_data["title"] = title
    except:
        pass

    # ---------------- Company ----------------
    try:
        company = header.locator(
            "a[data-wa-click='djv-job-company-profile-click']"
        ).inner_text().strip()
        if company:
            job_data["company"] = company
    except:
        pass

    # ---------------- Location + Posted ----------------
    try:
        meta = header.locator("span.text-sm").inner_text()
        parts = [p.strip() for p in meta.split("•")]

        if parts and parts[0]:
            job_data["location"] = parts[0]

        for p in parts:
            if "Posted" in p:
                posted_clean = p.replace("Posted", "").strip()
                if posted_clean:
                    job_data["posted"] = posted_clean
    except:
        pass

    # ---------------- Badges ----------------
    badges = header.locator("div.SeuiInfoBadge")
    for i in range(badges.count()):
        text = badges.nth(i).inner_text().strip()

        if not text:
            continue

        if "Full" in text:
            job_data["job_type"] = "Full Time"

        elif "Contract" in text:
            job_data["job_type"] = "Contract"

        if "$" in text:
            job_data["rate"] = text

    return job_data

def main():

    playwright = sync_playwright().start()

    browser = playwright.chromium.launch(
        headless=False,
        args=["--disable-blink-features=AutomationControlled"]
    )

    context = browser.new_context()
    base_page = context.new_page()

    base_page.goto(BASE_URL, timeout=60000)
    base_page.wait_for_load_state("networkidle")
    accept_cookies_if_present(base_page)

    ensure_location_united_states(base_page)

    for job_title in config["job_titles"]:

        print(f"\n[INFO] ===== Job title: {job_title} =====")

        # 🔥 Reset page completely
        base_page.goto(BASE_URL, timeout=60000)
        base_page.wait_for_load_state("networkidle")

        # Now search
        base_page.fill("input[placeholder*='Job title']", job_title)
        base_page.keyboard.press("Enter")
        base_page.wait_for_load_state("networkidle")
        ensure_location_united_states(base_page)
        base_page.keyboard.press("Enter")
        max_pages = config.get("scraper", {}).get("max_pages")

        # Apply filters AFTER search
        open_filters(base_page)
        apply_all_filters(base_page, config["filters"])

        base_page.wait_for_timeout(1500)

        # 3️⃣ Check if no results
        if base_page.locator("text=No results found").count() > 0 \
                or base_page.locator("text=0 results").count() > 0:

            print(f"[INFO] No jobs found for '{job_title}' — skipping sheet creation")
            continue

        page_number = 1
        seen_urls = set()

        while True:

            print(f"\n[INFO] Processing Page {page_number}")

            base_page.wait_for_selector("//div[@role='listitem']", timeout=20000)

            base_jobs = fetch_jobs_from_listitems(base_page)
            print(f"[INFO] Found {len(base_jobs)} jobs")

            for base in base_jobs:

                url = base["url"]

                if url in seen_urls:
                    continue
                seen_urls.add(url)

                detail_page = context.new_page()
                detail_page.goto(url, timeout=45000, wait_until="commit")
                detail_page.wait_for_load_state("domcontentloaded", timeout=20000)
                detail_page.wait_for_timeout(1500)

                details = extract_header_details(detail_page, base)
                append_job_to_excel(job_title, details)

                detail_page.close()

            # Stop if max_pages reached
            if max_pages and page_number >= max_pages:
                print("[INFO] Reached max_pages limit")
                break

            # Scroll to bottom to reveal pagination
            base_page.evaluate("window.scrollTo(0, document.body.scrollHeight)")
            base_page.wait_for_timeout(1000)

            # Locate Next arrow
            next_button = base_page.locator("//span[@aria-label='Next' and @role='link']")

            if next_button.count() == 0:
                print("[INFO] Next button not found — ending pagination")
                break

            # Check disabled state
            is_disabled = next_button.first.get_attribute("aria-disabled")

            if is_disabled == "true":
                print("[INFO] Next button disabled — last page reached")
                break

            print("[INFO] Moving to next page")

            next_button.first.click(force=True)

            base_page.wait_for_load_state("networkidle")
            base_page.wait_for_selector("//div[@role='listitem']", timeout=20000)

            page_number += 1


    browser.close()
    playwright.stop()


if __name__ == "__main__":
    main()
