import os
import requests
import yaml
from datetime import date, datetime, timedelta
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter


# ------------------------------------------------------
# Load config.yaml
# ------------------------------------------------------
def load_config(path="config.yaml"):
    with open(path, "r", encoding="utf-8") as f:
        return yaml.safe_load(f)


# ------------------------------------------------------
# Load sites list
# ------------------------------------------------------
def load_sites(path):
    with open(path, "r", encoding="utf-8") as f:
        return [line.strip() for line in f if line.strip()]


# ------------------------------------------------------
# Chunk list
# ------------------------------------------------------
def chunk_list(lst, size):
    for i in range(0, len(lst), size):
        yield lst[i:i+size]


# ------------------------------------------------------
# Build Google query
# ------------------------------------------------------
def build_query(job_title, keywords, regions, after_date, sites):
    keyword_part = " OR ".join([f'"{k}"' for k in keywords])
    region_part = " OR ".join([f'"{r}"' for r in regions])
    site_part = " OR ".join([f"site:{s}" for s in sites])

    query = (
        f'"{job_title}" ({keyword_part}) ({region_part}) '
        f'({site_part}) after:{after_date}'
    )
    return query


# ------------------------------------------------------
# SerpAPI Search
# ------------------------------------------------------
def serpapi_search(query, api_key, max_retries=5):
    url = "https://serpapi.com/search"
    params = {"engine": "google", "q": query, "num": "10", "api_key": api_key}

    for attempt in range(max_retries):
        try:
            r = requests.get(url, params=params, timeout=30)

            if r.status_code == 429:
                wait = 5 * (attempt + 1)
                print(f"[RATE LIMITED] 429 → Waiting {wait}s...")
                import time; time.sleep(wait)
                continue

            r.raise_for_status()
            return r.json()

        except Exception as e:
            print(f"[Error] attempt {attempt+1}/{max_retries}: {e}")
            import time; time.sleep(3 * (attempt + 1))

    print("[FATAL] SerpAPI failed after retries.")
    return None


# ------------------------------------------------------
# Extract Jobs
# ------------------------------------------------------
def extract_jobs(results):
    jobs = []
    if not results:
        return jobs

    for item in results:
        location = (
                item.get("location")
                or item.get("address", {}).get("locality")
                or "N/A"
        )

        job = {
            "title": item.get("title", "N/A"),
            "url": item.get("link", "N/A"),
            "snippet": item.get("snippet", "N/A"),
            "company": item.get("source", "N/A"),
            "location": location,
            "posted": item.get("date", "N/A")
        }
        jobs.append(job)

    return jobs


# ------------------------------------------------------
# Remove duplicates by URL
# ------------------------------------------------------
def remove_duplicates(jobs):
    seen = set()
    unique = []
    for job in jobs:
        url = job["url"].lower().strip()
        if url not in seen:
            seen.add(url)
            unique.append(job)
    return unique


# ------------------------------------------------------
# Format TXT Results
# ------------------------------------------------------
def format_results(job_title, jobs):
    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    text = f"JOB SEARCH RESULTS FOR: {job_title}\n"
    text += f"Search Date: {now}\n"
    text += "="*80 + "\n\n"

    if not jobs:
        text += "NO RESULTS FOUND\n"
        return text

    for idx, job in enumerate(jobs, start=1):
        text += f"--- Job {idx} ---\n"
        text += f"Title: {job['title']}\n"
        text += f"Company: {job['company']}\n"
        text += f"Location: {job['location']}\n"
        text += f"Posted: {job['posted']}\n"
        text += f"URL: {job['url']}\n"
        text += "Description:\n"
        text += f"  {job['snippet']}\n"
        text += "-"*80 + "\n\n"

    return text


# ------------------------------------------------------
# PROCESS ONE JOB TITLE — Lazy Excel Creation
# ------------------------------------------------------
def run_for_job(job_title, cfg, sites, serpapi_key, after_date,
                output_dir, queries_accumulator, wb, excel_path):

    if not job_title or not isinstance(job_title, str):
        print(f"Skipping invalid job title: {job_title}")
        return wb

    chunk_size = cfg["chunk_size"]
    keywords = cfg["keywords"]
    regions = cfg["regions"]

    print(f"\n==== PROCESSING: {job_title} ====\n")

    collected_jobs = []

    for idx, site_chunk in enumerate(chunk_list(sites, chunk_size), start=1):
        query = build_query(job_title, keywords, regions, after_date, site_chunk)
        queries_accumulator.append(f"[{job_title}] CHUNK {idx}: {query}")

        print(f"[{job_title}] Chunk {idx} → Searching...")

        serp = serpapi_search(query, serpapi_key)
        if not serp or "organic_results" not in serp:
            print(f"[{job_title}] Chunk {idx} → No results.")
            continue

        jobs = extract_jobs(serp["organic_results"])
        collected_jobs.extend(jobs)

        print(f"[{job_title}] Chunk {idx} → {len(jobs)} jobs")

    unique = remove_duplicates(collected_jobs)

    # ----------------------------------------------------
    # NO RESULTS → Skip TXT + Excel
    # ----------------------------------------------------
    if not unique:
        print(f"[{job_title}] NO RESULTS — skipping TXT + Excel.")
        return wb

    # ----------------------------------------------------
    # Write TXT file
    # ----------------------------------------------------
    job_slug = job_title.replace(" ", "")
    out_txt = os.path.join(output_dir, f"{job_slug}-results.txt")
    with open(out_txt, "w", encoding="utf-8") as f:
        f.write(format_results(job_title, unique))

    print(f"[{job_title}] Saved TXT → {out_txt}")

    # ----------------------------------------------------
    # Lazy Excel creation — only when results exist
    # ----------------------------------------------------
    if wb is None:
        wb = Workbook()
        wb.remove(wb.active)

    # ----------------------------------------------------
    # Add Excel sheet
    # ----------------------------------------------------
    sheet_name = job_slug[:31]
    if sheet_name in wb.sheetnames:
        del wb[sheet_name]

    ws = wb.create_sheet(title=sheet_name)
    headers = ["Title", "Company", "Location", "Posted", "URL", "Description"]
    ws.append(headers)

    for job in unique:
        ws.append([
            job.get("title", ""),
            job.get("company", ""),
            job.get("location", ""),
            job.get("posted", ""),
            job.get("url", ""),
            job.get("snippet", "")
        ])

    # Auto column width
    for col in ws.columns:
        max_len = max(len(str(cell.value)) for cell in col if cell.value)
        ws.column_dimensions[get_column_letter(col[0].column)].width = max_len + 2

    wb.save(excel_path)
    print(f"[{job_title}] Updated Excel → {excel_path}")

    return wb


# ------------------------------------------------------
# MAIN CONTROLLER — Lazy Excel Create
# ------------------------------------------------------
def main():

    BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    # Load config
    cfg = load_config(os.path.join(BASE_DIR, "config", "config.yaml"))

    serpapi_key = os.getenv("SERPAPI_KEY")
    if not serpapi_key:
        raise Exception("Missing SERPAPI_KEY environment variable.")

    # Load sites (absolute path)
    sites = load_sites(os.path.join(BASE_DIR, "config", "sites.txt"))

    # Force results inside /results folder
    output_root = os.path.join(BASE_DIR, "results")
    cfg["output_root"] = output_root

    after_date = (date.today() - timedelta(days=cfg["days_back"])).isoformat()
    today = date.today().isoformat()

    output_dir = os.path.join(cfg["output_root"], today)
    log_dir = os.path.join(output_dir, "log")

    os.makedirs(output_dir, exist_ok=True)
    os.makedirs(log_dir, exist_ok=True)

    excel_path = os.path.join(log_dir, "all-job-results.xlsx")

    wb = None
    if os.path.exists(excel_path):
        wb = load_workbook(excel_path)
        print("Loaded existing Excel workbook.")

    queries_log = []
    valid_titles = [jt for jt in cfg["job_titles"] if jt and isinstance(jt, str)]

    for job_title in valid_titles:
        wb = run_for_job(
            job_title, cfg, sites, serpapi_key, after_date,
            output_dir, queries_log, wb, excel_path
        )

    log_file = os.path.join(log_dir, "queries-and-results.txt")
    with open(log_file, "w", encoding="utf-8") as f:
        f.write("\n".join(queries_log))

    print(f"\nSaved queries → {log_file}")
    print("\nPROCESS COMPLETED\n")

if __name__ == "__main__":
    main()