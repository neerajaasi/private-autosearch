"""
Daily Job Scraper v2 — Direct ATS, Freshness-Aware
====================================================
Runs on your laptop. Hits direct ATS sources (Greenhouse + Lever JSON APIs +
JS-rendered Workday/Ashby/company sites via Playwright). Filters to your role
keywords + recency window. Outputs Excel.

Setup (one-time, ~2 min):
    pip install playwright requests openpyxl
    playwright install chromium

Run (every day, ~3-5 min):
    python daily_job_scraper.py

Output:
    job_report_YYYYMMDD_HHMM.xlsx (in same folder)

Customize:
    - ROLE_KEYWORDS       which roles to filter for (line ~30)
    - RECENCY_DAYS        how fresh a job must be to keep (line ~36)
    - GREENHOUSE_BOARDS   list of company slugs (line ~50)
    - LEVER_BOARDS        list of company slugs (line ~80)
    - JS_RENDERED_SITES   for Ashby/Workday/etc (line ~100)
"""

import asyncio
import json
import re
from datetime import datetime, timedelta, timezone
from pathlib import Path
from concurrent.futures import ThreadPoolExecutor, as_completed
import requests
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment


# ============================================================
# CONFIGURATION — Edit these to suit your job hunt
# ============================================================

# Roles to look for. Any keyword match (case-insensitive) keeps the job.
# Use phrases for precision; words to broaden.
ROLE_KEYWORDS = {
    "BA":    ["business analyst", "business systems analyst", "business operations analyst"],
    "QA":    ["qa engineer", "qa analyst", "quality assurance", "sdet", "test engineer", "test automation"],
    "DA":    ["data analyst", "analytics engineer", "bi analyst", "business intelligence analyst"],
    "Cyber": ["security analyst", "soc analyst", "security engineer", "cybersecurity",
              "information security", "iam engineer", "grc analyst", "penetration tester"],
}

# Only keep jobs posted within this many days. Set to 1 for "today only."
RECENCY_DAYS = 3

# Country/location filter. Empty list = all locations.
LOCATION_KEYWORDS = ["united states", "usa", "us-", "remote", "new york", "san francisco",
                     "seattle", "austin", "boston", "chicago", "atlanta", "denver",
                     "los angeles", "washington", "dallas", "houston"]

OUTPUT_FILE = f"job_report_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"

# ============================================================
# GREENHOUSE BOARDS — JSON API, super reliable
# Format: company slug from job-boards.greenhouse.io/<slug>
# Add / remove freely; finding more is easy: search Google for
# "site:job-boards.greenhouse.io <industry>"
# ============================================================
GREENHOUSE_BOARDS = [
    # Tech / SaaS
    "cloudflare", "reddit", "stripe", "instacart", "doordash", "robinhood",
    "elitetechnology", "fleetio", "purestorage", "taskrabbit", "jetbrains",
    # Healthcare / fintech
    "springhealth66", "smarterdx", "missionlane", "earnest", "merceradvisors",
    # Other
    "thenewyorktimes", "vestmark", "toastmastersinternational",
    "accenturefederalservices", "ctccampusboard",
]

# ============================================================
# LEVER BOARDS — JSON API, also reliable
# Format: company slug from jobs.lever.co/<slug>
# ============================================================
LEVER_BOARDS = [
    "h1", "fundera", "newsbreak", "ramp", "gusto",
    "attentivemobile", "mercury", "matterport", "klarna",
]

# ============================================================
# ASHBY BOARDS — JSON API (public posting endpoint)
# Format: company slug from jobs.ashbyhq.com/<slug>
# ============================================================
ASHBY_BOARDS = [
    "harvey", "virtahealth", "dandy", "sunnydata", "techtorch",
    "radiant-industries", "CARIAN", "brightwheel", "thatgamecompany",
    "jump-app", "hive.co", "pearlhealth", "ruby-labs", "airapps",
    "arb-interactive", "the-global-talent-co",
]

# ============================================================
# JS-RENDERED FALLBACK — uses Playwright for sites without an API
# Add Workday URLs (search results pages with all roles listed)
# Format: { "Company name": "Workday search URL" }
# ============================================================
JS_RENDERED_SITES = {
    "Prudential":         "https://pru.wd5.myworkdayjobs.com/Careers",
    "Amgen":              "https://amgen.wd1.myworkdayjobs.com/careers",
    "Western Union":      "https://westernunion.wd5.myworkdayjobs.com/WesternUnionJobs",
    "Capital One":        "https://capitalone.wd12.myworkdayjobs.com/en-US/Capital_One",
    "Marathon Petroleum": "https://mpc.wd1.myworkdayjobs.com/en-US/MPCCareers",
    "JLL":                "https://jll.wd1.myworkdayjobs.com/jllcareers",
}


# ============================================================
# CORE LOGIC — you generally don't need to edit below this line
# ============================================================

CUTOFF = datetime.now(timezone.utc) - timedelta(days=RECENCY_DAYS)
ALL_KEYWORDS = [(cat, kw.lower()) for cat, kws in ROLE_KEYWORDS.items() for kw in kws]


def match_role(title):
    """Return list of role categories this title matches."""
    t = title.lower()
    matched = []
    for cat, kw in ALL_KEYWORDS:
        if kw in t and cat not in matched:
            matched.append(cat)
    return matched


def match_location(loc):
    """Return True if location matches our filters (or no filter set)."""
    if not LOCATION_KEYWORDS:
        return True
    if not loc:
        return True  # missing location, don't exclude
    l = loc.lower()
    return any(kw in l for kw in LOCATION_KEYWORDS)


def parse_iso(s):
    """Parse ISO timestamp and return UTC datetime, or None if unparseable."""
    if not s:
        return None
    try:
        if s.endswith("Z"):
            s = s[:-1] + "+00:00"
        dt = datetime.fromisoformat(s)
        if dt.tzinfo is None:
            dt = dt.replace(tzinfo=timezone.utc)
        return dt.astimezone(timezone.utc)
    except (ValueError, TypeError):
        return None


# --------------------------------------------------
# Greenhouse: JSON API
# --------------------------------------------------
def fetch_greenhouse(slug):
    """Pull all jobs from a Greenhouse board via public JSON API."""
    url = f"https://boards-api.greenhouse.io/v1/boards/{slug}/jobs?content=true"
    out = []
    try:
        r = requests.get(url, timeout=15)
        if r.status_code != 200:
            return out, f"HTTP {r.status_code}"
        data = r.json()
        for job in data.get("jobs", []):
            title = job.get("title", "")
            cats = match_role(title)
            if not cats:
                continue
            loc = (job.get("location") or {}).get("name", "")
            if not match_location(loc):
                continue
            updated_at = parse_iso(job.get("updated_at"))
            if updated_at and updated_at < CUTOFF:
                continue
            out.append({
                "source":   "Greenhouse",
                "company":  slug,
                "title":    title,
                "location": loc,
                "posted":   updated_at.strftime("%Y-%m-%d") if updated_at else "Unknown",
                "url":      job.get("absolute_url", ""),
                "categories": ", ".join(cats),
            })
        return out, "OK"
    except Exception as e:
        return out, f"ERR: {str(e)[:60]}"


# --------------------------------------------------
# Lever: JSON API
# --------------------------------------------------
def fetch_lever(slug):
    url = f"https://api.lever.co/v0/postings/{slug}?mode=json"
    out = []
    try:
        r = requests.get(url, timeout=15)
        if r.status_code != 200:
            return out, f"HTTP {r.status_code}"
        data = r.json()
        for job in data:
            title = job.get("text", "")
            cats = match_role(title)
            if not cats:
                continue
            loc = ((job.get("categories") or {}).get("location", "") or "")
            if not match_location(loc):
                continue
            # Lever returns createdAt in milliseconds
            created_ms = job.get("createdAt", 0)
            posted_dt = datetime.fromtimestamp(created_ms / 1000, tz=timezone.utc) if created_ms else None
            if posted_dt and posted_dt < CUTOFF:
                continue
            out.append({
                "source":   "Lever",
                "company":  slug,
                "title":    title,
                "location": loc,
                "posted":   posted_dt.strftime("%Y-%m-%d") if posted_dt else "Unknown",
                "url":      job.get("hostedUrl", ""),
                "categories": ", ".join(cats),
            })
        return out, "OK"
    except Exception as e:
        return out, f"ERR: {str(e)[:60]}"


# --------------------------------------------------
# Ashby: JSON API
# --------------------------------------------------
def fetch_ashby(slug):
    """Ashby's public jobs feed."""
    url = f"https://api.ashbyhq.com/posting-api/job-board/{slug}"
    out = []
    try:
        r = requests.get(url, timeout=15)
        if r.status_code != 200:
            return out, f"HTTP {r.status_code}"
        data = r.json()
        for job in data.get("jobs", []):
            title = job.get("title", "")
            cats = match_role(title)
            if not cats:
                continue
            loc = job.get("locationName", "")
            if not match_location(loc):
                continue
            posted_dt = parse_iso(job.get("publishedAt") or job.get("updatedAt"))
            if posted_dt and posted_dt < CUTOFF:
                continue
            out.append({
                "source":   "Ashby",
                "company":  slug,
                "title":    title,
                "location": loc,
                "posted":   posted_dt.strftime("%Y-%m-%d") if posted_dt else "Unknown",
                "url":      job.get("jobUrl", ""),
                "categories": ", ".join(cats),
            })
        return out, "OK"
    except Exception as e:
        return out, f"ERR: {str(e)[:60]}"


# --------------------------------------------------
# Playwright fallback for JS-rendered career sites
# --------------------------------------------------
async def fetch_playwright(company, url, browser):
    """Best-effort scrape of a JS-rendered careers page."""
    out = []
    page = None
    try:
        page = await browser.new_page()
        await page.goto(url, timeout=30000, wait_until="domcontentloaded")
        await page.wait_for_timeout(4000)  # let JS settle

        # Try common job-link selectors
        selectors = [
            "a[data-automation-id='jobTitle']",  # Workday
            "a[href*='/job/']",
            "a[href*='/jobs/']",
            "h3 a",
            "li.css-1q2dra3 a",
        ]
        seen = set()
        for sel in selectors:
            try:
                links = await page.query_selector_all(sel)
                for link in links:
                    title = (await link.text_content() or "").strip()
                    if not title or len(title) < 5 or title in seen:
                        continue
                    cats = match_role(title)
                    if not cats:
                        continue
                    seen.add(title)
                    href = await link.get_attribute("href") or ""
                    if href and not href.startswith("http"):
                        from urllib.parse import urljoin
                        href = urljoin(url, href)
                    out.append({
                        "source":     "Playwright",
                        "company":    company,
                        "title":      title,
                        "location":   "(see job page)",
                        "posted":     "(unknown — Workday/JS site)",
                        "url":        href,
                        "categories": ", ".join(cats),
                    })
                if out:
                    break  # one selector worked; move on
            except Exception:
                continue
        return out, "OK" if out else "NO_MATCHES"
    except Exception as e:
        return out, f"ERR: {str(e)[:60]}"
    finally:
        if page:
            await page.close()


async def run_playwright():
    """Run all JS-rendered fallback scrapes in one browser instance."""
    try:
        from playwright.async_api import async_playwright
    except ImportError:
        print("⚠️  Playwright not installed. Skipping JS-rendered sites.")
        print("    To enable: pip install playwright && playwright install chromium")
        return [], {}

    all_results = []
    statuses = {}
    async with async_playwright() as p:
        browser = await p.chromium.launch(headless=True)
        try:
            tasks = [fetch_playwright(c, u, browser) for c, u in JS_RENDERED_SITES.items()]
            results = await asyncio.gather(*tasks, return_exceptions=True)
            for (company, _), result in zip(JS_RENDERED_SITES.items(), results):
                if isinstance(result, Exception):
                    statuses[f"Playwright/{company}"] = f"EXCEPTION: {str(result)[:60]}"
                    continue
                jobs, status = result
                statuses[f"Playwright/{company}"] = f"{status} ({len(jobs)} found)"
                all_results.extend(jobs)
        finally:
            await browser.close()
    return all_results, statuses


# --------------------------------------------------
# Main
# --------------------------------------------------
def main():
    print(f"Daily Job Scraper — {datetime.now().strftime('%Y-%m-%d %H:%M')}")
    print(f"Cutoff: jobs posted on or after {CUTOFF.strftime('%Y-%m-%d')}")
    print(f"Roles: {', '.join(ROLE_KEYWORDS.keys())}")
    print()

    all_jobs = []
    source_status = {}

    # ----- Phase 1: API-based fetches in parallel -----
    print("Phase 1: Fetching API-based sources...")
    with ThreadPoolExecutor(max_workers=12) as pool:
        future_to_label = {}
        for slug in GREENHOUSE_BOARDS:
            f = pool.submit(fetch_greenhouse, slug)
            future_to_label[f] = f"Greenhouse/{slug}"
        for slug in LEVER_BOARDS:
            f = pool.submit(fetch_lever, slug)
            future_to_label[f] = f"Lever/{slug}"
        for slug in ASHBY_BOARDS:
            f = pool.submit(fetch_ashby, slug)
            future_to_label[f] = f"Ashby/{slug}"

        for fut in as_completed(future_to_label):
            label = future_to_label[fut]
            try:
                jobs, status = fut.result()
                source_status[label] = f"{status} ({len(jobs)} kept)"
                all_jobs.extend(jobs)
                if jobs:
                    print(f"  ✓ {label:40s} {len(jobs)} jobs")
            except Exception as e:
                source_status[label] = f"EXCEPTION: {str(e)[:60]}"

    # ----- Phase 2: JS-rendered scraping -----
    print(f"\nPhase 2: Scraping JS-rendered sites ({len(JS_RENDERED_SITES)} sites)...")
    pw_jobs, pw_status = asyncio.run(run_playwright())
    source_status.update(pw_status)
    all_jobs.extend(pw_jobs)
    if pw_jobs:
        print(f"  ✓ Playwright: {len(pw_jobs)} jobs from JS sites")

    # ----- Dedupe by URL -----
    seen_urls = set()
    deduped = []
    for j in all_jobs:
        if j["url"] in seen_urls:
            continue
        seen_urls.add(j["url"])
        deduped.append(j)

    print(f"\nResults: {len(all_jobs)} jobs found, {len(deduped)} after dedup")

    # ----- Build Excel report -----
    wb = Workbook()
    ws = wb.active
    ws.title = "Jobs Found"

    headers = ["#", "Category", "Source", "Company", "Job Title", "Location", "Posted", "Apply Link"]
    for c_idx, h in enumerate(headers, start=1):
        c = ws.cell(row=1, column=c_idx, value=h)
        c.font = Font(name="Arial", bold=True, color="FFFFFF", size=11)
        c.fill = PatternFill("solid", start_color="1F4E78")
        c.alignment = Alignment(horizontal="center", vertical="center")

    cat_colors = {
        "BA":    "C6EFCE",
        "QA":    "FFC7CE",
        "DA":    "DDEBF7",
        "Cyber": "E4D4F4",
    }

    # Sort: most recent first, then by category
    def sort_key(j):
        try:
            return (datetime.strptime(j["posted"], "%Y-%m-%d").date(), j["categories"])
        except (ValueError, TypeError):
            return (datetime(1970, 1, 1).date(), j["categories"])

    deduped.sort(key=sort_key, reverse=True)

    for i, j in enumerate(deduped, start=2):
        primary_cat = j["categories"].split(",")[0].strip()
        fill = PatternFill("solid", start_color=cat_colors.get(primary_cat, "FFFFFF"))
        row = [i - 1, j["categories"], j["source"], j["company"], j["title"],
               j["location"], j["posted"], j["url"]]
        for c_idx, val in enumerate(row, start=1):
            c = ws.cell(row=i, column=c_idx, value=val)
            c.font = Font(name="Arial", size=10)
            c.fill = fill
            c.alignment = Alignment(vertical="top", wrap_text=True)
            if c_idx == 8:
                c.font = Font(name="Arial", size=10, color="0563C1", underline="single")
                if val:
                    c.hyperlink = val

    widths = [5, 14, 14, 22, 42, 28, 14, 50]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[chr(64 + i)].width = w
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    # Source status sheet
    ws2 = wb.create_sheet("Source Status")
    ws2.cell(row=1, column=1, value="Source").font = Font(bold=True)
    ws2.cell(row=1, column=2, value="Status").font = Font(bold=True)
    for i, (label, status) in enumerate(sorted(source_status.items()), start=2):
        ws2.cell(row=i, column=1, value=label)
        ws2.cell(row=i, column=2, value=status)
    ws2.column_dimensions["A"].width = 50
    ws2.column_dimensions["B"].width = 50

    wb.save(OUTPUT_FILE)
    print(f"\n✅ Saved: {OUTPUT_FILE}")

    # Summary by category
    from collections import Counter
    cats = Counter()
    for j in deduped:
        for c in j["categories"].split(","):
            cats[c.strip()] += 1
    print("\nBreakdown by role:")
    for c, n in cats.most_common():
        print(f"  {c:8s} {n}")


if __name__ == "__main__":
    main()
