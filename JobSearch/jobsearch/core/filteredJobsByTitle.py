import os
import requests
import yaml
from datetime import date, datetime, timedelta
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter


# ------------------------------------------------------
# Load config.yaml
# ------------------------------------------------------
def load_config(path):
    with open(path, "r", encoding="utf-8") as f:
        return yaml.safe_load(f)


# ------------------------------------------------------
# Normalize Posted-At Date ("1 day ago", "3 hours ago")
# ------------------------------------------------------
def normalize_posted(text):
    text = text.lower().strip()
    today = date.today()

    if "hour" in text or "minute" in text:
        return today

    if "day" in text:
        days = int(text.split()[0])
        return today - timedelta(days=days)

    if "week" in text:
        weeks = int(text.split()[0])
        return today - timedelta(days=weeks * 7)

    return None


# ------------------------------------------------------
# Return all job-type keywords based on comma-separated modes
# ------------------------------------------------------
def get_selected_jobtype_keywords(cfg):
    modes_str = cfg["job_type"].get("modes", "").lower().strip()
    if not modes_str:
        return []

    modes = [m.strip() for m in modes_str.split(",") if m.strip()]
    selected_keywords = []

    for mode in modes:
        kw_list = cfg["job_type"]["include_keywords"].get(mode, [])
        selected_keywords.extend([kw.lower() for kw in kw_list])

    return selected_keywords


# ------------------------------------------------------
# Build Google Jobs Query
# ------------------------------------------------------
def build_google_jobs_query(job_title, keywords, regions, after_date, cfg):
    keyword_part = " OR ".join([f'"{k}"' for k in keywords])
    region_part = " OR ".join([f'"{r}"' for r in regions])

    type_keywords = get_selected_jobtype_keywords(cfg)
    type_part = " OR ".join([f'"{kw}"' for kw in type_keywords]) if type_keywords else ""
    posted_filter = (
        f'"posted after {after_date}" OR '
        f'"posted on {after_date}" OR '
        f'"{after_date}"'
    )

    query = (
        f'"{job_title}" ({keyword_part}) ({region_part}) '
        f'({type_part}) ({posted_filter})'
    )

    return query


# ------------------------------------------------------
# SerpAPI Google Jobs Search
# ------------------------------------------------------
def serpapi_jobs_search(query, api_key, max_retries=5):
    url = "https://serpapi.com/search"
    params = {"engine": "google_jobs", "q": query, "api_key": api_key}

    for attempt in range(max_retries):
        try:
            r = requests.get(url, params=params, timeout=30)

            if r.status_code == 429:
                wait = 5 * (attempt + 1)
                print(f"[RATE LIMITED] 429 → Waiting {wait}s...")
                import time; time.sleep(wait)
                continue

            r.raise_for_status()
            return r.json()

        except Exception as e:
            print(f"[ERROR] SerpAPI attempt {attempt+1}: {e}")
            import time; time.sleep(3 * (attempt + 1))

    print("[FATAL] SerpAPI failed after max retries.")
    return None


# ------------------------------------------------------
# Extract + Filter Results
# ------------------------------------------------------
def extract_google_jobs(results, cfg, after_date):
    jobs = []
    if not results or "jobs_results" not in results:
        return jobs

    after_dt = datetime.strptime(after_date, "%Y-%m-%d").date()
    mode_keywords = get_selected_jobtype_keywords(cfg)

    for item in results["jobs_results"]:
        desc = item.get("description", "").lower()
        schedule_type = item.get("detected_extensions", {}).get("schedule_type", "").lower()

        posted = item.get("detected_extensions", {}).get("posted_at", "N/A")
        posted_date = normalize_posted(posted) or after_dt

        if posted_date < after_dt:
            continue

        if mode_keywords:
            if not any(kw in desc or kw in schedule_type for kw in mode_keywords):
                continue

        job_url = "N/A"
        if "apply_options" in item and item["apply_options"]:
            job_url = item["apply_options"][0].get("link", "N/A")

        job = {
            "title": item.get("title", "N/A"),
            "company": item.get("company_name", "N/A"),
            "location": item.get("location", "N/A"),
            "posted": posted,
            "salary": item.get("detected_extensions", {}).get("salary", "N/A"),
            "jobtype": schedule_type,
            "url": job_url,
            "snippet": item.get("description", "N/A")
        }
        jobs.append(job)

    return jobs


# ------------------------------------------------------
# Remove duplicates by URL
# ------------------------------------------------------
def remove_duplicates(jobs):
    seen = set()
    unique = []
    for job in jobs:
        url = job["url"].lower().strip()
        if url not in seen:
            seen.add(url)
            unique.append(job)
    return unique


# ------------------------------------------------------
# Format TXT Output
# ------------------------------------------------------
def format_results(job_title, jobs):
    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    text = f"JOB SEARCH RESULTS FOR: {job_title}\n"
    text += f"Search Date: {now}\n"
    text += "=" * 80 + "\n\n"

    if not jobs:
        text += "NO RESULTS FOUND\n"
        return text

    for idx, job in enumerate(jobs, start=1):
        text += f"--- Job {idx} ---\n"
        text += f"Title: {job['title']}\n"
        text += f"Company: {job['company']}\n"
        text += f"Location: {job['location']}\n"
        text += f"Posted: {job['posted']}\n"
        text += f"Salary: {job['salary']}\n"
        text += f"JobType: {job['jobtype']}\n"
        text += f"URL: {job['url']}\n"
        text += "-" * 80 + "\n\n"

    return text


# ------------------------------------------------------
# Process individual job title
# ------------------------------------------------------
def run_for_job(job_title, cfg, serpapi_key, after_date,
                output_dir, queries_accumulator, wb, excel_path):

    print(f"\n==== PROCESSING: {job_title} ====\n")

    keywords = cfg["keywords"]
    regions = cfg["regions"]

    query = build_google_jobs_query(job_title, keywords, regions, after_date, cfg)
    queries_accumulator.append(f"[{job_title}] {query}")

    serp = serpapi_jobs_search(query, serpapi_key)
    jobs = extract_google_jobs(serp, cfg, after_date)
    unique = remove_duplicates(jobs)

    print(f"[{job_title}] Found {len(unique)} results")

    # Always save TXT
    job_slug = job_title.replace(" ", "")
    out_txt = os.path.join(output_dir, f"{job_slug}-results.txt")
    if not unique:
        print(f"[{job_title}] No results → skipping TXT creation.")
    else:
        with open(out_txt, "w", encoding="utf-8") as f:
            f.write(format_results(job_title, unique))
        print(f"[{job_title}] Saved TXT → {out_txt}")

    # ------------------------------------------------------
    #  CRITICAL FIX: SKIP SHEET CREATION IF NO RESULTS
    # ------------------------------------------------------
    if not unique:
        print(f"[{job_title}] No results → skipping Excel sheet creation.")
        return wb

    # Initialize workbook ONLY when we get first non-empty result
    if wb is None:
        wb = Workbook()
        ws = wb.active
        ws.title = job_slug[:31]   # name sheet immediately
    else:
        # Remove existing sheet if present
        sheet_name = job_slug[:31]
        if sheet_name in wb.sheetnames:
            del wb[sheet_name]
        ws = wb.create_sheet(title=sheet_name)

    # Write headers
    headers = ["Title", "Company", "Location", "Posted", "Salary", "JobType", "URL", "Description"]
    ws.append(headers)

    # Write data
    for job in unique:
        ws.append([
            job["title"], job["company"], job["location"], job["posted"],
            job["salary"], job["jobtype"], job["url"], job["snippet"]
        ])

    # Auto-fit columns
    for col in ws.columns:
        max_len = max(len(str(cell.value)) for cell in col if cell.value)
        ws.column_dimensions[get_column_letter(col[0].column)].width = max_len + 2

    # Save the workbook
    wb.save(excel_path)
    print(f"[{job_title}] Excel Updated → {excel_path}")

    return wb


# ------------------------------------------------------
# MAIN ENTRY POINT
# ------------------------------------------------------
def main():
    BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))

    cfg_path = os.path.join(BASE_DIR, "config", "config.yaml")
    cfg = load_config(cfg_path)

    serpapi_key = os.getenv("SERPAPI_KEY")
    if not serpapi_key:
        raise Exception("Missing SERPAPI_KEY environment variable.")

    today = date.today().isoformat()

    output_root = os.path.join(BASE_DIR, "results", "filteredjob", today)
    log_dir = os.path.join(output_root, "log")

    os.makedirs(output_root, exist_ok=True)
    os.makedirs(log_dir, exist_ok=True)

    excel_path = os.path.join(log_dir, "filtered-job-results.xlsx")

    # Do NOT create workbook here → only created when first result exists
    wb = load_workbook(excel_path) if os.path.exists(excel_path) else None

    after_date = (date.today() - timedelta(days=cfg["days_back"])).isoformat()
    queries_log = []

    for job_title in cfg["job_titles"]:
        wb = run_for_job(
            job_title, cfg, serpapi_key, after_date,
            output_root, queries_log, wb, excel_path
        )

    # Save queries log
    with open(os.path.join(log_dir, "queries.txt"), "w", encoding="utf-8") as f:
        f.write("\n".join(queries_log))

    print("\nALL JOB SEARCHES COMPLETED.\n")


if __name__ == "__main__":
    main()
