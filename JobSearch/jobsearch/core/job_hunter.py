"""
End-to-End Job Hunter — v4 (Workday file reader + 6 roles)
==============================================================
Reads Workday URLs from `workday_sites.txt` (one URL per line).
Fetches BA, QA, DA, SQL DBA, Cyber, and BDA roles posted in last 24 hours.

Setup (one-time, ~3 min):
    pip install playwright requests openpyxl
    playwright install chromium

Run:
    python job_hunter.py

Required file in same folder:
    workday_sites.txt    — one Workday URL per line

Output:
    job_hunter_report_<timestamp>.xlsx
"""

import asyncio
import json
import re
import time
from datetime import datetime, timedelta, timezone
from pathlib import Path
from urllib.parse import urlparse, parse_qs
from concurrent.futures import ThreadPoolExecutor, as_completed
from collections import Counter
import requests
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment


# ============================================================
# CONFIGURATION
# ============================================================

RECENCY_HOURS = 24  # strict 24h

ROLE_KEYWORDS = {
    "BA": [
        "business analyst", "business systems analyst", "business operations analyst",
        "sr business analyst", "senior business analyst", "lead business analyst",
        "principal business analyst",
    ],
    "QA": [
        "qa engineer", "qa analyst", "quality assurance", "sdet",
        "test engineer", "test automation", "automation engineer",
        "software test", "qa lead", "quality engineer",
    ],
    "DA": [
        "data analyst", "analytics engineer", "bi analyst",
        "sr data analyst", "senior data analyst",
        "lead data analyst", "principal data analyst",
    ],
    "SQL_DBA": [
        "sql administrator", "sql server administrator", "database administrator",
        "sql dba", " dba", "sql server dba", "oracle dba", "postgres dba",
        "database engineer", "sql developer", "database admin",
    ],
    "Cyber": [
        "security analyst", "soc analyst", "security engineer", "cybersecurity",
        "cyber security", "information security", "infosec", "iam engineer",
        "iam analyst", "grc analyst", "grc engineer", "penetration tester",
        "security operations", "threat analyst", "incident response",
        "security architect", "vulnerability analyst", "application security",
    ],
    "BDA": [
        "business data analyst", "business intelligence analyst",
        "bi developer", "data and analytics", "data insights analyst",
        "reporting analyst", "data reporting", "insights analyst",
    ],
}

LOCATION_KEYWORDS = [
    "united states", "usa", "us-", "us ", "remote",
    "new york", "san francisco", "seattle", "austin", "boston",
    "chicago", "atlanta", "denver", "los angeles", "washington",
    "dallas", "houston", "ny,", "ca,", "tx,", "ma,", "wa,",
]

DISCOVERY_ENABLED = True
CACHE_FILE = "company_cache.json"
WORKDAY_FILE = str(Path(__file__).parent / "workday_sites.txt")
OUTPUT_FILE = f"job_hunter_report_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"


# ============================================================
# SEED LISTS — Greenhouse, Lever, Ashby
# ============================================================

SEED_GREENHOUSE = [
    "cloudflare", "reddit", "stripe", "instacart", "doordash", "robinhood",
    "elitetechnology", "fleetio", "purestorage", "taskrabbit", "jetbrains",
    "springhealth66", "smarterdx", "missionlane", "earnest", "merceradvisors",
    "thenewyorktimes", "vestmark", "toastmastersinternational",
    "accenturefederalservices", "ctccampusboard",
    "airbnb", "anthropic", "openai", "snowflake", "databricks", "figma",
    "notion", "plaid", "asana", "atlassian", "vercel", "linear",
    "discord", "segment", "twilio", "duolingo", "wise", "remitly",
    "rippling", "deel", "samsara", "mongodb", "elastic", "hashicorp",
    "thumbtack", "zillow", "udemy", "coursera", "khanacademy",
]

SEED_LEVER = [
    "h1", "fundera", "newsbreak", "ramp", "gusto",
    "attentivemobile", "mercury", "matterport", "klarna",
    "wealthsimple", "scaleai", "alpaca", "alloy", "remote",
]

SEED_ASHBY = [
    "harvey", "virtahealth", "dandy", "sunnydata", "techtorch",
    "radiant-industries", "brightwheel", "thatgamecompany",
    "jump-app", "hive.co", "pearlhealth", "ruby-labs", "airapps",
    "perplexity", "cursor", "supabase",
]

EXCLUDE_SLUGS = {"jobgether", "remoterocketship"}


# ============================================================
# WORKDAY URL PARSER + FILE LOADER
# ============================================================

def parse_workday_url(url):
    """
    Parse a Workday URL into (display_name, base_url, tenant_id, site_id).

    Examples it handles:
      https://salesforce.wd12.myworkdayjobs.com/External_Career_Site
      https://nvidia.wd5.myworkdayjobs.com/NVIDIAExternalCareerSite/
      https://nvidia.wd5.myworkdayjobs.com/en-US/NVIDIAExternalCareerSite

    Returns (None, None, None, None) if URL can't be parsed.
    """
    url = url.strip().rstrip("/")
    if not url or not url.startswith("http"):
        return None, None, None, None

    parsed = urlparse(url)
    host = parsed.netloc
    if not host or "myworkdayjobs.com" not in host:
        return None, None, None, None

    m = re.match(r"^([a-z0-9_-]+)\.wd[0-9]+\.myworkdayjobs\.com$", host, re.I)
    if not m:
        return None, None, None, None
    tenant_id = m.group(1)

    path = parsed.path.strip("/")
    parts = [p for p in path.split("/") if p]
    if not parts:
        return None, None, None, None
    # Drop optional /en-US/ locale prefix
    if len(parts) > 1 and re.match(r"^[a-z]{2}-[A-Z]{2}$", parts[0]):
        site_id = parts[1]
    else:
        site_id = parts[0]

    base_url = f"https://{host}"
    display = tenant_id.replace("-", " ").replace("_", " ").title()

    return display, base_url, tenant_id, site_id


def load_workday_sites(path=WORKDAY_FILE):
    """Read workday_sites.txt and return a list of parsed tuples.
    Skips blank lines and lines starting with #."""
    p = Path(path)
    if not p.exists():
        print(f"⚠️  {path} not found. Workday section will be empty.")
        print(f"   Create the file with one URL per line, e.g.:")
        print(f"     https://salesforce.wd12.myworkdayjobs.com/External_Career_Site")
        return []

    tenants = []
    skipped = []
    seen = set()
    with open(p) as f:
        for lineno, raw in enumerate(f, 1):
            line = raw.strip()
            if not line or line.startswith("#"):
                continue
            display, base, tenant, site = parse_workday_url(line)
            if not tenant:
                skipped.append((lineno, line))
                continue
            key = (base, tenant, site)
            if key in seen:
                continue
            seen.add(key)
            tenants.append((display, base, tenant, site))

    if skipped:
        print(f"⚠️  Skipped {len(skipped)} unparseable lines in {path}:")
        for lineno, line in skipped[:5]:
            print(f"     L{lineno}: {line[:80]}")
        if len(skipped) > 5:
            print(f"     ... and {len(skipped) - 5} more")

    return tenants


# ============================================================
# DISCOVERY (Greenhouse / Lever / Ashby only)
# ============================================================

DISCOVERY_QUERIES = [
    ("greenhouse", '"business analyst" site:job-boards.greenhouse.io'),
    ("greenhouse", '"data analyst" site:job-boards.greenhouse.io'),
    ("greenhouse", '"qa engineer" site:job-boards.greenhouse.io'),
    ("greenhouse", '"database administrator" site:job-boards.greenhouse.io'),
    ("greenhouse", '"security analyst" site:job-boards.greenhouse.io'),
    ("greenhouse", '"senior business analyst" site:boards.greenhouse.io'),
    ("greenhouse", '"data analyst" site:boards.greenhouse.io'),
    ("lever",      '"business analyst" site:jobs.lever.co'),
    ("lever",      '"data analyst" site:jobs.lever.co'),
    ("lever",      '"qa engineer" site:jobs.lever.co'),
    ("lever",      '"security engineer" site:jobs.lever.co'),
    ("ashby",      '"business analyst" site:jobs.ashbyhq.com'),
    ("ashby",      '"data analyst" site:jobs.ashbyhq.com'),
    ("ashby",      '"qa engineer" site:jobs.ashbyhq.com'),
    ("ashby",      '"security analyst" site:jobs.ashbyhq.com'),
]

SLUG_PATTERNS = {
    "greenhouse": [
        re.compile(r"job-boards\.greenhouse\.io/([a-z0-9_-]+)/", re.I),
        re.compile(r"boards\.greenhouse\.io/([a-z0-9_-]+)/", re.I),
    ],
    "lever": [re.compile(r"jobs\.lever\.co/([a-z0-9_.-]+)/", re.I)],
    "ashby": [
        re.compile(r"jobs\.ashbyhq\.com/([a-zA-Z0-9_.-]+)/", re.I),
        re.compile(r"jobs\.ashbyhq\.com/([a-zA-Z0-9_.-]+)$", re.I),
    ],
}


def extract_slug(url, platform):
    for pat in SLUG_PATTERNS[platform]:
        m = pat.search(url)
        if m:
            slug = m.group(1).lower().strip()
            if slug in ("api", "v1", "v0", "search", "jobs", "embed", "_next"):
                return None
            return slug
    return None


async def google_search_one(page, query, max_results=30):
    urls = []
    encoded = query.replace(" ", "+").replace('"', "%22")
    search_url = f"https://www.google.com/search?q={encoded}&num={max_results}"
    try:
        await page.goto(search_url, timeout=20000, wait_until="domcontentloaded")
        await page.wait_for_timeout(2500)
        content = await page.content()
        if "unusual traffic" in content.lower() or "/sorry/" in page.url:
            return urls, "CAPTCHA"
        anchors = await page.query_selector_all("a[href]")
        for a in anchors:
            href = await a.get_attribute("href") or ""
            if href.startswith("/url?"):
                qs = parse_qs(urlparse(href).query)
                if "q" in qs:
                    urls.append(qs["q"][0])
            elif href.startswith("http") and "google.com" not in href:
                urls.append(href)
        return urls, "OK"
    except Exception as e:
        return urls, f"ERR: {str(e)[:60]}"


async def discover_companies():
    found = {"greenhouse": set(), "lever": set(), "ashby": set()}
    try:
        from playwright.async_api import async_playwright
    except ImportError:
        print("⚠️  Playwright not installed. Skipping discovery.")
        return found, "PLAYWRIGHT_MISSING"

    print(f"🔍 Discovery: running {len(DISCOVERY_QUERIES)} Google searches...")
    captcha_hits = 0
    async with async_playwright() as p:
        browser = await p.chromium.launch(
            headless=True,
            args=["--disable-blink-features=AutomationControlled"],
        )
        context = await browser.new_context(
            user_agent="Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                       "AppleWebKit/537.36 (KHTML, like Gecko) "
                       "Chrome/124.0.0.0 Safari/537.36",
            viewport={"width": 1366, "height": 900},
        )
        page = await context.new_page()
        for platform, query in DISCOVERY_QUERIES:
            urls, status = await google_search_one(page, query)
            if status == "CAPTCHA":
                captcha_hits += 1
                if captcha_hits >= 3:
                    print("    ⚠️  Hit Google CAPTCHA 3 times. Stopping discovery.")
                    break
            new_slugs = {extract_slug(u, platform) for u in urls}
            new_slugs.discard(None)
            new_slugs -= EXCLUDE_SLUGS
            found[platform].update(new_slugs)
            print(f"    {platform:11s} '{query[:45]:45s}'  {status:10s}  {len(new_slugs)} found")
            await asyncio.sleep(2.5)
        await browser.close()
    return found, "OK" if captcha_hits < 3 else "PARTIAL_CAPTCHA"


# ============================================================
# CACHE
# ============================================================

def load_cache():
    if not Path(CACHE_FILE).exists():
        return {"greenhouse": [], "lever": [], "ashby": []}
    try:
        with open(CACHE_FILE) as f:
            return json.load(f)
    except Exception:
        return {"greenhouse": [], "lever": [], "ashby": []}


def save_cache(merged):
    with open(CACHE_FILE, "w") as f:
        json.dump({k: sorted(v) for k, v in merged.items()}, f, indent=2)


def merge_lists(seed_gh, seed_lever, seed_ashby, discovered, cache):
    gh = set(seed_gh) | set(cache.get("greenhouse", []))
    lever = set(seed_lever) | set(cache.get("lever", []))
    ashby = set(seed_ashby) | set(cache.get("ashby", []))
    if discovered:
        gh.update(discovered.get("greenhouse", set()))
        lever.update(discovered.get("lever", set()))
        ashby.update(discovered.get("ashby", set()))
    gh -= EXCLUDE_SLUGS
    lever -= EXCLUDE_SLUGS
    ashby -= EXCLUDE_SLUGS
    return {"greenhouse": gh, "lever": lever, "ashby": ashby}


# ============================================================
# MATCHING
# ============================================================

CUTOFF = datetime.now(timezone.utc) - timedelta(hours=RECENCY_HOURS)
ALL_KEYWORDS = [(cat, kw.lower()) for cat, kws in ROLE_KEYWORDS.items() for kw in kws]
# Sort longer keywords first so "senior business analyst" matches before "business analyst"
ALL_KEYWORDS.sort(key=lambda x: -len(x[1]))


def match_role(title):
    """Returns list of role categories the title matches."""
    t = title.lower()
    matched = []
    for cat, kw in ALL_KEYWORDS:
        if kw in t and cat not in matched:
            matched.append(cat)
    return matched


def match_location(loc):
    if not LOCATION_KEYWORDS:
        return True
    if not loc:
        return True
    l = loc.lower()
    return any(kw in l for kw in LOCATION_KEYWORDS)


def parse_iso(s):
    if not s:
        return None
    try:
        if isinstance(s, (int, float)):
            return datetime.fromtimestamp(s / 1000 if s > 1e10 else s, tz=timezone.utc)
        if s.endswith("Z"):
            s = s[:-1] + "+00:00"
        dt = datetime.fromisoformat(s)
        if dt.tzinfo is None:
            dt = dt.replace(tzinfo=timezone.utc)
        return dt.astimezone(timezone.utc)
    except (ValueError, TypeError):
        return None


def parse_workday_posted(text):
    """Workday returns 'Posted Today', 'Posted Yesterday', 'Posted N Days Ago',
    'Posted N+ Days Ago'.

    STRICT 24h: only 'Today' counts.
    'Yesterday' = 24h ago, which is at the cutoff and gets dropped.
    """
    if not text:
        return None
    t = text.lower()
    now = datetime.now(timezone.utc)
    if "today" in t:
        return now
    if "yesterday" in t:
        return now - timedelta(hours=24)
    m = re.search(r"posted (\d+)\+?\s*days?\s*ago", t)
    if m:
        return now - timedelta(days=int(m.group(1)))
    m = re.search(r"posted (\d+)\+?\s*hours?\s*ago", t)
    if m:
        return now - timedelta(hours=int(m.group(1)))
    return None


def hours_ago_str(posted_dt):
    if not posted_dt:
        return ""
    delta = datetime.now(timezone.utc) - posted_dt
    h = int(delta.total_seconds() // 3600)
    if h < 24:
        return f"{h}h ago"
    return f"{h // 24}d ago"


# ============================================================
# FETCHERS
# ============================================================

def fetch_greenhouse(slug):
    url = f"https://boards-api.greenhouse.io/v1/boards/{slug}/jobs?content=true"
    out = []
    try:
        r = requests.get(url, timeout=12)
        if r.status_code != 200:
            return out, f"HTTP {r.status_code}"
        for job in r.json().get("jobs", []):
            title = job.get("title", "")
            cats = match_role(title)
            if not cats:
                continue
            loc = (job.get("location") or {}).get("name", "")
            if not match_location(loc):
                continue
            posted = parse_iso(job.get("updated_at"))
            if posted and posted < CUTOFF:
                continue
            out.append({
                "source": "Greenhouse", "company": slug, "title": title,
                "location": loc,
                "posted": posted.strftime("%Y-%m-%d %H:%M") if posted else "Unknown",
                "hours_ago": hours_ago_str(posted),
                "url": job.get("absolute_url", ""),
                "categories": ", ".join(cats),
                "_sort_dt": posted or datetime(1970, 1, 1, tzinfo=timezone.utc),
            })
        return out, "OK"
    except Exception as e:
        return out, f"ERR: {str(e)[:60]}"


def fetch_lever(slug):
    url = f"https://api.lever.co/v0/postings/{slug}?mode=json"
    out = []
    try:
        r = requests.get(url, timeout=12)
        if r.status_code != 200:
            return out, f"HTTP {r.status_code}"
        for job in r.json():
            title = job.get("text", "")
            cats = match_role(title)
            if not cats:
                continue
            loc = ((job.get("categories") or {}).get("location", "") or "")
            if not match_location(loc):
                continue
            created_ms = job.get("createdAt", 0)
            posted = parse_iso(created_ms) if created_ms else None
            if posted and posted < CUTOFF:
                continue
            out.append({
                "source": "Lever", "company": slug, "title": title,
                "location": loc,
                "posted": posted.strftime("%Y-%m-%d %H:%M") if posted else "Unknown",
                "hours_ago": hours_ago_str(posted),
                "url": job.get("hostedUrl", ""),
                "categories": ", ".join(cats),
                "_sort_dt": posted or datetime(1970, 1, 1, tzinfo=timezone.utc),
            })
        return out, "OK"
    except Exception as e:
        return out, f"ERR: {str(e)[:60]}"


def fetch_ashby(slug):
    url = f"https://api.ashbyhq.com/posting-api/job-board/{slug}"
    out = []
    try:
        r = requests.get(url, timeout=12)
        if r.status_code != 200:
            return out, f"HTTP {r.status_code}"
        for job in r.json().get("jobs", []):
            title = job.get("title", "")
            cats = match_role(title)
            if not cats:
                continue
            loc = job.get("locationName", "")
            if not match_location(loc):
                continue
            posted = parse_iso(job.get("publishedAt") or job.get("updatedAt"))
            if posted and posted < CUTOFF:
                continue
            out.append({
                "source": "Ashby", "company": slug, "title": title,
                "location": loc,
                "posted": posted.strftime("%Y-%m-%d %H:%M") if posted else "Unknown",
                "hours_ago": hours_ago_str(posted),
                "url": job.get("jobUrl", ""),
                "categories": ", ".join(cats),
                "_sort_dt": posted or datetime(1970, 1, 1, tzinfo=timezone.utc),
            })
        return out, "OK"
    except Exception as e:
        return out, f"ERR: {str(e)[:60]}"


def fetch_workday(display_name, base_url, tenant_id, site_id):
    """Hit Workday's internal job-search API. Pages through results."""
    api_url = f"{base_url}/wday/cxs/{tenant_id}/{site_id}/jobs"
    out = []
    headers = {
        "Content-Type": "application/json",
        "Accept": "application/json",
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                      "AppleWebKit/537.36 (KHTML, like Gecko) "
                      "Chrome/124.0.0.0 Safari/537.36",
    }

    offset = 0
    page_size = 20
    max_pages = 8
    pages_fetched = 0

    try:
        while pages_fetched < max_pages:
            payload = {
                "limit": page_size, "offset": offset,
                "searchText": "", "appliedFacets": {},
            }
            try:
                r = requests.post(api_url, json=payload, headers=headers, timeout=15)
            except requests.exceptions.SSLError:
                r = requests.post(api_url, json=payload, headers=headers,
                                  timeout=15, verify=False)
            if r.status_code != 200:
                if pages_fetched == 0:
                    return out, f"HTTP {r.status_code}"
                break

            data = r.json()
            jobs = data.get("jobPostings", [])
            if not jobs:
                break

            stale_count = 0
            for job in jobs:
                title = job.get("title", "")
                cats = match_role(title)
                if not cats:
                    continue
                loc = job.get("locationsText", "")
                if not match_location(loc):
                    continue
                posted = parse_workday_posted(job.get("postedOn", ""))
                if posted and posted < CUTOFF:
                    stale_count += 1
                    continue
                ext_path = job.get("externalPath", "")
                full_url = f"{base_url}{ext_path}" if ext_path else base_url
                out.append({
                    "source": "Workday", "company": display_name, "title": title,
                    "location": loc,
                    "posted": (posted.strftime("%Y-%m-%d") if posted else
                               job.get("postedOn", "Unknown")),
                    "hours_ago": hours_ago_str(posted) if posted else "",
                    "url": full_url,
                    "categories": ", ".join(cats),
                    "_sort_dt": posted or datetime(1970, 1, 1, tzinfo=timezone.utc),
                })

            # Workday returns jobs roughly post-date-desc.
            # If most jobs on page are stale, no point paging further.
            if stale_count >= page_size * 0.7:
                break

            pages_fetched += 1
            offset += page_size
            total = data.get("total", 0)
            if offset >= total:
                break
            time.sleep(0.3)

        return out, "OK"
    except Exception as e:
        return out, f"ERR: {str(e)[:60]}"


# ============================================================
# EXCEL OUTPUT
# ============================================================

CAT_COLORS = {
    "BA":      "C6EFCE",  # light green
    "QA":      "FFC7CE",  # light pink
    "DA":      "DDEBF7",  # light blue
    "SQL_DBA": "FFE699",  # light orange/yellow
    "Cyber":   "E4D4F4",  # light purple
    "BDA":     "D9F2D0",  # mint
}

CAT_DESCRIPTIONS = {
    "BA":      "Business Analyst",
    "QA":      "Quality Assurance / Test",
    "DA":      "Data Analyst",
    "SQL_DBA": "SQL / Database Administrator",
    "Cyber":   "Cybersecurity",
    "BDA":     "Business Data Analyst / BI",
}


def write_excel(jobs, source_status, run_meta):
    wb = Workbook()
    ws = wb.active
    ws.title = "Fresh Jobs"

    role_list = "/".join(ROLE_KEYWORDS.keys())
    ws["A1"] = (f"Job Hunt Report — {datetime.now().strftime('%Y-%m-%d %H:%M')} — "
                f"{role_list} roles posted in last {RECENCY_HOURS} hours")
    ws["A1"].font = Font(bold=True, size=13, color="1F4E78")
    ws.merge_cells("A1:I1")

    ws["A2"] = (f"Companies scanned: {run_meta['total_companies']} | "
                f"Greenhouse: {run_meta['gh']}, Lever: {run_meta['lever']}, "
                f"Ashby: {run_meta['ashby']}, Workday: {run_meta['workday']} | "
                f"Discovery: {run_meta['discovery_status']}")
    ws["A2"].font = Font(italic=True, size=10, color="555555")
    ws.merge_cells("A2:I2")

    headers = ["#", "Category", "Source", "Company", "Job Title", "Location",
               "Posted", "Time Ago", "Apply Link"]
    for c_idx, h in enumerate(headers, start=1):
        c = ws.cell(row=4, column=c_idx, value=h)
        c.font = Font(bold=True, color="FFFFFF", size=11)
        c.fill = PatternFill("solid", start_color="1F4E78")
        c.alignment = Alignment(horizontal="center", vertical="center")

    jobs.sort(key=lambda j: j["_sort_dt"], reverse=True)

    for i, j in enumerate(jobs, start=5):
        primary_cat = j["categories"].split(",")[0].strip()
        fill = PatternFill("solid", start_color=CAT_COLORS.get(primary_cat, "FFFFFF"))
        row_data = [i - 4, j["categories"], j["source"], j["company"],
                    j["title"], j["location"], j["posted"], j["hours_ago"], j["url"]]
        for c_idx, val in enumerate(row_data, start=1):
            c = ws.cell(row=i, column=c_idx, value=val)
            c.font = Font(size=10)
            c.fill = fill
            c.alignment = Alignment(vertical="top", wrap_text=True)
            if c_idx == 9 and val:
                c.font = Font(size=10, color="0563C1", underline="single")
                c.hyperlink = val

    widths = [4, 14, 12, 22, 42, 28, 17, 11, 50]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[chr(64 + i)].width = w
    ws.freeze_panes = "A5"
    if jobs:
        ws.auto_filter.ref = f"A4:I{4 + len(jobs)}"

    # Source status sheet
    ws2 = wb.create_sheet("Source Status")
    ws2.cell(row=1, column=1, value="Source").font = Font(bold=True)
    ws2.cell(row=1, column=2, value="Status").font = Font(bold=True)
    for i, (label, status) in enumerate(sorted(source_status.items()), start=2):
        ws2.cell(row=i, column=1, value=label)
        ws2.cell(row=i, column=2, value=status)
    ws2.column_dimensions["A"].width = 50
    ws2.column_dimensions["B"].width = 50
    ws2.freeze_panes = "A2"

    # Legend sheet
    ws3 = wb.create_sheet("Legend")
    ws3["A1"] = "Role Categories"
    ws3["A1"].font = Font(bold=True, size=14, color="1F4E78")
    ws3.cell(row=3, column=1, value="Code").font = Font(bold=True)
    ws3.cell(row=3, column=2, value="Description").font = Font(bold=True)
    ws3.cell(row=3, column=3, value="Keywords matched").font = Font(bold=True)
    for i, (cat, kws) in enumerate(ROLE_KEYWORDS.items(), start=4):
        c = ws3.cell(row=i, column=1, value=cat)
        c.font = Font(bold=True, size=11)
        c.fill = PatternFill("solid", start_color=CAT_COLORS.get(cat, "FFFFFF"))
        ws3.cell(row=i, column=2, value=CAT_DESCRIPTIONS.get(cat, ""))
        ws3.cell(row=i, column=3, value=", ".join(kws))
        ws3.cell(row=i, column=3).alignment = Alignment(wrap_text=True, vertical="top")
    ws3.column_dimensions["A"].width = 12
    ws3.column_dimensions["B"].width = 30
    ws3.column_dimensions["C"].width = 80

    wb.save(OUTPUT_FILE)


# ============================================================
# MAIN
# ============================================================

def main():
    start = time.time()
    print(f"\n{'='*60}")
    print(f"Job Hunter v4 — End-to-End Pipeline")
    print(f"{'='*60}")
    print(f"Time:           {datetime.now().strftime('%Y-%m-%d %H:%M')}")
    print(f"Recency window: last {RECENCY_HOURS} hours (STRICT)")
    print(f"Roles:          {', '.join(ROLE_KEYWORDS.keys())}")
    print()

    # PHASE 1: Load Workday URLs from file
    print(f"📂 Loading Workday URLs from {WORKDAY_FILE}...")
    workday_list = load_workday_sites()
    print(f"   Loaded {len(workday_list)} Workday tenants\n")

    # PHASE 2: Discovery
    discovered = None
    discovery_status = "skipped"
    if DISCOVERY_ENABLED:
        try:
            discovered, discovery_status = asyncio.run(discover_companies())
            new_total = sum(len(s) for s in discovered.values())
            print(f"  → Discovery returned {new_total} candidate slugs\n")
        except Exception as e:
            discovery_status = f"FAILED: {str(e)[:80]}"
            print(f"  ⚠️  Discovery failed: {e}\n")

    # PHASE 3: Merge
    cache = load_cache()
    merged = merge_lists(SEED_GREENHOUSE, SEED_LEVER, SEED_ASHBY, discovered, cache)
    save_cache(merged)
    gh_list = sorted(merged["greenhouse"])
    lever_list = sorted(merged["lever"])
    ashby_list = sorted(merged["ashby"])
    total = len(gh_list) + len(lever_list) + len(ashby_list) + len(workday_list)
    print(f"📊 Companies to scan: {total}")
    print(f"   Greenhouse: {len(gh_list)} | Lever: {len(lever_list)} | "
          f"Ashby: {len(ashby_list)} | Workday: {len(workday_list)}\n")

    # PHASE 4: Fetch
    print(f"⚡ Fetching jobs from all {total} companies in parallel...")
    all_jobs = []
    source_status = {}
    with ThreadPoolExecutor(max_workers=25) as pool:
        futures = {}
        for slug in gh_list:
            futures[pool.submit(fetch_greenhouse, slug)] = f"Greenhouse/{slug}"
        for slug in lever_list:
            futures[pool.submit(fetch_lever, slug)] = f"Lever/{slug}"
        for slug in ashby_list:
            futures[pool.submit(fetch_ashby, slug)] = f"Ashby/{slug}"
        for display, base, tenant, site in workday_list:
            futures[pool.submit(fetch_workday, display, base, tenant, site)] = f"Workday/{display}"

        for fut in as_completed(futures):
            label = futures[fut]
            try:
                jobs, status = fut.result()
                source_status[label] = f"{status} ({len(jobs)} matches)"
                all_jobs.extend(jobs)
            except Exception as e:
                source_status[label] = f"EXCEPTION: {str(e)[:60]}"

    # PHASE 5: Dedupe
    seen_urls = set()
    deduped = []
    for j in all_jobs:
        if not j["url"] or j["url"] in seen_urls:
            continue
        seen_urls.add(j["url"])
        deduped.append(j)

    # PHASE 6: Output
    elapsed = int(time.time() - start)
    run_meta = {
        "total_companies": total,
        "gh": len(gh_list), "lever": len(lever_list),
        "ashby": len(ashby_list), "workday": len(workday_list),
        "discovery_status": discovery_status,
    }
    write_excel(deduped, source_status, run_meta)

    print(f"\n{'='*60}")
    print(f"✅ DONE — completed in {elapsed}s")
    print(f"{'='*60}")
    print(f"Jobs found:     {len(all_jobs)}")
    print(f"After dedup:    {len(deduped)}")
    print(f"Output file:    {OUTPUT_FILE}\n")

    cats = Counter()
    by_source = Counter()
    for j in deduped:
        for c in j["categories"].split(","):
            cats[c.strip()] += 1
        by_source[j["source"]] += 1
    if cats:
        print("By role:")
        for c, n in cats.most_common():
            print(f"  {c:10s} {n}")
        print("\nBy source:")
        for s, n in by_source.most_common():
            print(f"  {s:12s} {n}")
    else:
        print(f"⚠️  No fresh jobs in last {RECENCY_HOURS}h.")
        print(f"    Try setting RECENCY_HOURS = 48 or 72 to verify the script works.")
    print()


if __name__ == "__main__":
    main()
