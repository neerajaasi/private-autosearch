#!/usr/bin/env python3
"""
main.py — Multi-site Job Scraper
==================================
Reads config.yaml for job titles, date filter, location.
Reads sites.txt for which sites to scrape.
Outputs Excel with sheets per job title.

Usage:
  python main.py
  python main.py --no-headless
  python main.py --sites roberthalf --titles "Data Analyst"
"""

import argparse
import logging
import sys
import time
from datetime import datetime
from pathlib import Path

import yaml
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

SCRIPT_DIR = Path(__file__).resolve().parent
sys.path.insert(0, str(SCRIPT_DIR))

from driver_factory import create_driver
from sites import REGISTRY

CONFIG_FILE = SCRIPT_DIR / "config.yaml"
SITES_FILE = SCRIPT_DIR / "sites.txt"

COLUMNS = ["Company", "Title", "Location", "Salary", "Job Type", "Work Type", "Posted Date", "JD", "URL"]


def setup_logging(verbose: bool = False):
    logging.basicConfig(
        level=logging.DEBUG if verbose else logging.INFO,
        format="%(asctime)s │ %(levelname)-7s │ %(message)s",
        datefmt="%H:%M:%S",
    )


def load_config() -> dict:
    if not CONFIG_FILE.exists():
        logging.error(f"Config not found: {CONFIG_FILE}")
        sys.exit(1)
    with open(CONFIG_FILE) as f:
        return yaml.safe_load(f)


def load_sites() -> list[str]:
    if not SITES_FILE.exists():
        logging.error(f"Sites file not found: {SITES_FILE}")
        sys.exit(1)
    sites = []
    with open(SITES_FILE) as f:
        for line in f:
            line = line.strip()
            if line and not line.startswith("#"):
                name = line.lower()
                if name in REGISTRY:
                    sites.append(name)
                else:
                    logging.warning(f"Unknown site '{name}'. Available: {', '.join(REGISTRY.keys())}")
    return sites


def sanitize_sheet_name(name: str) -> str:
    clean = name.replace("/", "-").replace("\\", "-").replace("*", "").replace("?", "")
    clean = clean.replace("[", "").replace("]", "").replace(":", "-")
    return clean[:31]


def style_header(ws, num_cols: int):
    header_font = Font(name="Arial", bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill("solid", fgColor="2F5496")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for col in range(1, num_cols + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align

    col_widths = {
        "Company": 16, "Title": 40, "Location": 22, "Salary": 28,
        "Job Type": 18, "Work Type": 12, "Posted Date": 14, "JD": 80, "URL": 55,
    }
    for i, col_name in enumerate(COLUMNS, 1):
        ws.column_dimensions[get_column_letter(i)].width = col_widths.get(col_name, 20)

    ws.freeze_panes = "A2"


def write_jobs_to_sheet(ws, jobs: list[dict]):
    data_font = Font(name="Arial", size=10)
    wrap_align = Alignment(vertical="top", wrap_text=True)
    url_font = Font(name="Arial", size=10, color="0563C1", underline="single")
    alt_fill = PatternFill("solid", fgColor="F2F2F2")

    for row_idx, job in enumerate(jobs, 2):
        values = [
            job.get("company", ""),
            job.get("title", ""),
            job.get("location", ""),
            job.get("salary", ""),
            job.get("job_type", ""),
            job.get("work_type", ""),
            job.get("posted_date", ""),
            (job.get("jd", "") or "")[:500],
            job.get("url", ""),
        ]
        for col_idx, val in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.font = data_font
            cell.alignment = wrap_align
            if row_idx % 2 == 0:
                cell.fill = alt_fill

        url_val = job.get("url", "")
        if url_val:
            url_cell = ws.cell(row=row_idx, column=len(COLUMNS))
            url_cell.font = url_font
            try:
                url_cell.hyperlink = url_val
            except Exception:
                pass


def build_excel(all_results: dict, config: dict, output_path: Path):
    wb = Workbook()

    # Summary sheet (first)
    ws_summary = wb.active
    ws_summary.title = "Summary"
    ws_summary.append(["Job Title", "Total Jobs", "Sites with Results"])
    style_header(ws_summary, 3)
    ws_summary.column_dimensions["A"].width = 30
    ws_summary.column_dimensions["B"].width = 15
    ws_summary.column_dimensions["C"].width = 45

    for title, jobs in all_results.items():
        sites_found = sorted(set(j.get("company", "") for j in jobs))
        ws_summary.append([title, len(jobs), ", ".join(sites_found)])

    ws_summary.append([])
    total_row = ws_summary.max_row + 1
    ws_summary.cell(row=total_row, column=1, value="TOTAL").font = Font(name="Arial", bold=True, size=11)
    ws_summary.cell(row=total_row, column=2, value=sum(len(v) for v in all_results.values())).font = Font(name="Arial", bold=True, size=11)

    # All Jobs sheet
    ws_all = wb.create_sheet("All Jobs")
    ws_all.append(COLUMNS)
    style_header(ws_all, len(COLUMNS))
    all_flat = []
    for jobs in all_results.values():
        all_flat.extend(jobs)
    write_jobs_to_sheet(ws_all, all_flat)

    # One sheet per job title
    for title, jobs in all_results.items():
        sheet_name = sanitize_sheet_name(title)
        ws = wb.create_sheet(title=sheet_name)
        ws.append(COLUMNS)
        style_header(ws, len(COLUMNS))
        write_jobs_to_sheet(ws, jobs)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(output_path))
    logging.info(f"📊 Excel saved: {output_path}")


def main():
    parser = argparse.ArgumentParser(description="Multi-site Job Scraper")
    parser.add_argument("--no-headless", action="store_true", help="Show browser")
    parser.add_argument("--sites", default=None, help="Comma-separated sites (overrides sites.txt)")
    parser.add_argument("--titles", default=None, help="Comma-separated titles (overrides config.yaml)")
    parser.add_argument("--date-filter", default=None, help="Date filter (overrides config.yaml)")
    parser.add_argument("-v", "--verbose", action="store_true")
    args = parser.parse_args()

    setup_logging(args.verbose)
    logger = logging.getLogger(__name__)

    config = load_config()
    titles = [t.strip() for t in args.titles.split(",")] if args.titles else config["titles"]
    location = config.get("location", "united states")
    date_filter = args.date_filter or config.get("date_filter", "Past 24 hours")
    output_cfg = config.get("output", {})
    output_dir = SCRIPT_DIR / output_cfg.get("directory", "output")
    output_file = output_dir / output_cfg.get("filename", "job_results.xlsx")

    if args.sites:
        site_names = [s.strip().lower() for s in args.sites.split(",")]
    else:
        site_names = load_sites()

    if not site_names:
        logger.error("No sites configured.")
        sys.exit(1)

    print()
    print("═" * 60)
    print("  JOB SCRAPER")
    print(f"  Titles:      {', '.join(titles)}")
    print(f"  Sites:       {', '.join(site_names)}")
    print(f"  Location:    {location}")
    print(f"  Date filter: {date_filter}")
    print(f"  Output:      {output_file}")
    print("═" * 60)

    all_results: dict[str, list[dict]] = {t: [] for t in titles}

    for title in titles:
        print(f"\n{'─' * 60}")
        print(f"  🔍 Searching: {title}")
        print(f"{'─' * 60}")

        for site_name in site_names:
            scraper_cls = REGISTRY[site_name]
            driver = create_driver(headless=not args.no_headless)

            try:
                scraper = scraper_cls(driver, date_filter=date_filter)
                jobs = scraper.search_jobs(title, location)

                for job in jobs:
                    job["search_title"] = title

                all_results[title].extend(jobs)

                if jobs:
                    logger.info(f"[{site_name}] ✅ {len(jobs)} jobs for '{title}'")
                    for i, j in enumerate(jobs[:5], 1):
                        t = j["title"][:45]
                        loc = j.get("location", "")[:18]
                        sal = j.get("salary", "")[:20]
                        print(f"    {i}. {t:<45} {loc:<20} {sal}")
                    if len(jobs) > 5:
                        print(f"    ... +{len(jobs) - 5} more")
                else:
                    logger.warning(f"[{site_name}] ⚠ 0 jobs for '{title}'")

            except Exception as e:
                logger.error(f"[{site_name}] ❌ Error for '{title}': {e}")
                try:
                    err_dir = output_dir
                    err_dir.mkdir(parents=True, exist_ok=True)
                    driver.save_screenshot(str(err_dir / f"error_{site_name}_{title.replace(' ', '_')}.png"))
                except Exception:
                    pass
            finally:
                driver.quit()

            time.sleep(2)

    # Build Excel
    build_excel(all_results, config, output_file)

    # Final summary
    total = sum(len(v) for v in all_results.values())
    print(f"\n{'═' * 60}")
    print(f"  ✅ DONE — {total} total jobs")
    print(f"  📊 {output_file}")
    print()
    for title, jobs in all_results.items():
        companies = sorted(set(j.get("company", "?") for j in jobs))
        print(f"    {title:<30} {len(jobs):>4} jobs  ({', '.join(companies) or 'none'})")
    print(f"{'═' * 60}\n")


if __name__ == "__main__":
    main()
